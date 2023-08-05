"""Flatfile-related utilities"""
#%%
import os
import chardet
import pandas as pd
from openpyxl import Workbook
import gzip

#%%

def detect_encoding(file_path):
    """Attempt to determine the encoding of a file located at the provided file path"""
    # Check if the file path ends with '.gz' to identify gzipped files
    if file_path.endswith('.gz'):
        # Open the gzipped file in binary mode
        with gzip.open(file_path, 'rb') as f:
            # Read the first 2000 rows
            content = b''.join([f.readline() for _ in range(2000)])
    else:
        # Open the file in binary mode to prevent any decoding errors
        with open(file_path, 'rb') as f:
            # Read the first 2000 rows
            content = b''.join([f.readline() for _ in range(2000)])

    # Determine the encoding of the content
    result = chardet.detect(content)
    encoding = result['encoding']

    # If the detected encoding is ASCII, read the entire file to confirm the encoding
    if encoding.lower() == 'ascii':
        if file_path.endswith('.gz'):
            # Open the gzipped file in binary mode
            with gzip.open(file_path, 'rb') as f:
                content = f.read()
        else:
            # Open the file in binary mode to prevent any decoding errors
            with open(file_path, 'rb') as f:
                content = f.read()

        # Determine the encoding of the entire file
        result = chardet.detect(content)
        encoding = result['encoding']

    return encoding

def extract_file_extension(filename, last_only=False):
    extension_found = False
    extension = ''
    exclusions = ['.data','.script','.file','.image']
    if os.path.isabs(filename) or os.path.sep in filename:
        filename = os.path.basename(filename)
    
    if '.' not in filename:
        return 'Invalid filename'#, extension_found
    
    if last_only:
        extension = os.path.splitext(filename)[1]
        extension_found = True
        return extension#, extension_found
    
    parts = filename.split('.')
    
    if len(parts) == 2 and (len(parts[0]) == 0 or len(parts[1]) == 0):
        return 'No Extension'#, extension_found
    
    working_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(working_dir, 'references', 'known_file_extensions.txt')

    with open(file_path, 'r') as file:
        # extensions = [ext[1:] if ext.startswith('.') else ext for ext in file.read().splitlines()]
        extensions = [ext[1:] if ext.startswith('.') else ext for ext in file.read().splitlines() if ext not in exclusions]

    match_string = None
    
    for part in parts[1:]:
        if part in extensions:
            match_string = part
            break

    if match_string is None:
        return 'Unknown Extension'#, extension_found
    
    match_index = filename.find('.' + match_string)
    extension = filename[match_index:]

    if extension:
        extension_found = True
        return extension#, extension_found
    else:
        return 'Unknown Extension'#, extension_found

def analyze_dataframe(df):
    """Analyze distinct values in a dataframe"""
    analysis_dict = {}

    for col in df.columns:
        column_dict = {}

        # Add column header
        column_dict['Column'] = col
        distinct_value_count = df[col].nunique()
        distinct_value_count_string = str(format(distinct_value_count, ','))
        column_dict['Distinct Values Count'] = str(distinct_value_count)

        # Check column data type
        col_dtype = df[col].dtype
        if pd.api.types.is_string_dtype(col_dtype):
            # For string columns, calculate the maximum length
            max_size = df[col].str.len().max()
            max_size_string = str(int(max_size)) if not pd.isna(max_size) else ''
            column_dict['Max Size'] = f'Max Length: {max_size_string}'
            column_dict['Type'] = 'String'
        elif pd.api.types.is_numeric_dtype(col_dtype):
            if pd.api.types.is_integer_dtype(col_dtype):
                # For integer columns, convert the maximum value to int
                max_value = df[col].max()
                max_value_string = str(int(max_value)) if not pd.isna(max_value) else ''
                column_dict['Max Size'] = f'Max Value: {max_value_string}'
            else:
                # For other numeric columns (float), store the maximum value as is
                max_value_string = str(df[col].max())
                column_dict['Max Size'] = f'Max Value: {max_value_string}'
            column_dict['Type'] = 'Numeric'
        elif pd.api.types.is_datetime64_any_dtype(col_dtype):
            # For date or timestamp columns
            max_date = df[col].max()
            max_date_string = str(max_date) if not pd.isna(max_date) else ''
            column_dict['Max Size'] = f'Max Value: {max_date_string}'
            column_dict['Type'] = 'Date/Time'

        # Get distinct values and their counts
        col_values = df[col].copy()
        if col_values.dtype == float:
            col_values = col_values.apply(lambda x: "{:.10f}".format(x).rstrip('0').rstrip('.') if not pd.isna(x) else '<NULL>')
        col_values = col_values.fillna("<NULL>")
        value_counts = col_values.value_counts(dropna=False).replace({pd.NA: "<NULL>", pd.NaT: "<NULL>", "nan": "<NULL>"})
        if len(value_counts) > 50:
            # If there are more than 50 distinct values, store top 50 and "More than 50 distinct values"
            top_50_values = value_counts.nlargest(50).to_dict()
            # top_50_values['More than 50 distinct values'] = str(len(value_counts) - 50)
            # column_dict['Distinct Values'] = {str(k): str(format(int(v), ',')) if v != '' else str(v) for k, v in top_50_values.items()}
            column_dict['Distinct Values'] = {('{:.10f}'.format(k) if isinstance(k, float) else str(k)): str(v) if v != '' else str(v) for k, v in top_50_values.items()}
            # top_50_values['More than 50 distinct values'] = f'Distinct Values: {distinct_value_count_string}'
            column_dict['Distinct Values']['More than 50 distinct values'] = f'Distinct Values: {distinct_value_count_string}'

        else:
            # column_dict['Distinct Values'] = {str(k): str(format(int(v), ',')) if v != '' else str(v) for k, v in value_counts.items()}
            column_dict['Distinct Values'] = {('{:.10f}'.format(k) if isinstance(k, float) else str(k)): str(v) if v != '' else str(v) for k, v in value_counts.items()}

        # Add the column dictionary to the analysis dictionary
        analysis_dict[col] = column_dict

    # Create a workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Initialize the start column index
    start_column_index = 1

    # Iterate over each column in the analysis dictionary
    for col, col_dict in analysis_dict.items():
        # Write the "Column" and "Max Size" to the Excel file
        ws.cell(row=1, column=start_column_index, value=col_dict['Column'])
        if 'Max Size' in col_dict:
            ws.cell(row=1, column=start_column_index + 1, value=col_dict['Max Size'])

        ws.cell(row=2, column=start_column_index, value='Distinct Values')
        ws.cell(row=2, column=start_column_index + 1, value='Occurrences')

        # Write the "Distinct Values" to the Excel file
        distinct_values = col_dict['Distinct Values']
        for row_num, (key, value) in enumerate(distinct_values.items(), start=3):
            ws.cell(row=row_num, column=start_column_index, value=key)
            ws.cell(row=row_num, column=start_column_index + 1, value=value)

        # Increment the start column index for the next iteration
        start_column_index += 3

    ws.freeze_panes = 'A3'
    
    df_results = pd.DataFrame(wb.active.values)

    return df_results

def check_primary_key_fields(df, field_list, print_results=True):
    """Accepts a data frame and a list of field to determine if there are duplicates"""
    
    if not set(field_list).issubset(df.columns):
        missing_fields = set(field_list) - set(df.columns)
        missing_fields_string = ', '.join(f'"{field}"' for field in missing_fields)
        print(f'The following fields are missing in the DataFrame: {missing_fields_string}')
        return None, None, None
    
    is_unique = not df.duplicated(subset=field_list, keep=False).any()
    
    for field in field_list:
        no_nulls = not df[field].isnull().any()

    field_names = ', '.join(f'"{field}"' for field in field_list)

    sample_duplicate_rows = pd.DataFrame()
    sample_null_rows = pd.DataFrame()

    if is_unique:
        if print_results:
            print(f'The field(s) {field_names} have no duplicate values.')
    else:
        duplicate_rows = df[df.duplicated(subset=field_list, keep=False)]
        duplicate_first_row_values = duplicate_rows[field_list].head(1).values.tolist()[0]
        matching_rows = df.copy()
        
        for field, value in zip(field_list, duplicate_first_row_values):
            matching_rows = matching_rows[matching_rows[field] == value]
        
        matching_rows.sort_values(by=field_list, inplace=True)
        sample_duplicate_rows = matching_rows.head(2)
        
        if print_results:
            print(f'The field(s) {field_names} have duplicate values.')
            print("Sample:")
            print(df.columns.to_list())
            for _, row in sample_duplicate_rows.iterrows():
                print(row.values.tolist())

    if no_nulls:
        if print_results:
            print(f'The field(s) {field_names} have no NULL values.')
    else:
        for field in field_list:
            if df[field].isnull().any():
                null_row_index = df[df[field].isnull()].index[0]
                if null_row_index not in sample_null_rows.index.tolist():
                    sample_null_rows = pd.concat([sample_null_rows, pd.DataFrame([df.loc[null_row_index].values], columns=df.columns)])
                
        if print_results:
            print(f'The field(s) {field_names} have NULL values.')
            print("Sample:")
            print(df.columns.to_list())
            sample_null_rows.fillna("<NULL>", inplace=True)
            for _, row in sample_null_rows.iterrows():
                print(row.values.tolist())
    
    sample_issue_rows = pd.concat([sample_duplicate_rows, sample_null_rows], ignore_index=True)
    
    return is_unique, no_nulls, sample_issue_rows

#%%

if __name__ == '__main__':
    # filename = 'Testfile.txt.gz'
    # print(f'{filename}\n{extract_file_extension(filename)}\n\n')

    # filename = 'Test.file.txt.gz'
    # print(f'{filename}\n{extract_file_extension(filename)}\n\n')

    filename = 'my.data.csv.gz'
    print(f'{filename}\n{extract_file_extension(filename)}\n\n')

    filename = 'my.report.docx.template'
    print(f'{filename}\n{extract_file_extension(filename)}\n\n')

    filename = 'my.archive.tar.gz'
    print(f'{filename}\n{extract_file_extension(filename)}\n\n')

    filename = 'my.image.jpg.thumbnail'
    print(f'{filename}\n{extract_file_extension(filename)}\n\n')

    filename = 'my.script.py.txt'
    print(f'{filename}\n{extract_file_extension(filename, True)}\n\n')

    # filename = 'My.test.file.backup.txt.gz'
    # print(f'{filename}\n{extract_file_extension(filename)}\n\n')

    # filename = r'C:\test_folder\My.test.file.backup.txt.gz'
    # print(f'{filename}\n{extract_file_extension(filename)}\n\n')

    # filename = r'C:\test_folder'
    # print(f'{filename}\n{extract_file_extension(filename)}\n\n')

    # filename = 'C:\\test_folder\\My.test.file.backup.txt.gz'
    # print(f'{filename}\n{extract_file_extension(filename)}\n\n')

    pass

    #%%