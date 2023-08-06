import json
import requests
from requests.exceptions import RequestException
import re
import time
import random
import os
import logging
from lxml import etree
import sys
print(sys.path)
sys.path.append(".")
import time
import jenkins
from openpyxl import load_workbook
from pyocs.pyocs_software import PyocsSoftware
from pyocs import pyocs_software
from pyocs.pyocs_demand import PyocsDemand
from pyocs.pyocs_request import PyocsRequest
from pyocs import pyocs_login
from pyocs.pyocs_list import PyocsList
from pyocs.pyocs_request import PyocsRequest
import logging
from urllib.parse import urlencode
from lxml import etree
from colorama import Fore, Back, Style
from pyocs.pyocs_jenkins import PyocsJenkins

import smtplib
from email import encoders
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import parseaddr, formataddr
from pyocs import pyocs_email

import platform
#from pyocs import faas_api

import json
import base64

from_addr = 'linxiangna@cvte.com'
password = 'Lin13690439782'
smtp_server = 'smtp.exmail.qq.com'

new_ocs_url = "http://ocs.gz.cvte.cn/tv/Tasks/index/ContractTask//SearchId:2091062/range:all"
find_old_keliaohao_ocs_url = "http://ocs.gz.cvte.cn/tv/Tasks/index/range:all/"
auto_build_url = "http://tvci.gz.cvte.cn/job/yangpingbu/job/Sample_Compile/build"
reuse_stock_software_url = "https://tvci.gz.cvte.cn/job/Utility_Tools/job/reuse_stock_software/build"
_ocs_base_link = 'https://ocs.gz.cvte.cn'
_ocs_find_old_sw_url = _ocs_base_link + "/tv/pop/selectors/common_view/Controller:FirmwareAttachments/" \
                                            "Parameters:accountId/input_id:old_fw_id/input_name:old_fw_name/inputtype:1"
customer_name ="金品"
customer_order = "01-202005-132"
customer_liaohao="R41-63S2T2B00000A1C"
#member = "林祥纳 陈潮雄 陈嘉艺"
member = "林祥纳"
ocs_status="待录入需求"
#list_xpath = '//td[@class="Task_col_id"]|//td[@class="Task_col_sw_user_id"]|//td[@class="Task_col_status"]|//td[@class="Task_col_subject"]/a|//td[@class="Task_col_rd_dept_id"]'
list_xpath = '//td[@class="Task_col_id"]|//td[@class="Task_col_sw_user_id"]|//td[@class="Task_col_status"]|//td[@class="Task_col_subject"]/a|//td[@class="Task_col_rd_dept_id"]|//td[@class="Contract_col_account_prod_name"]'

#用vscode运行时，用这个
dirPath = os.getcwd()+'/customers/customer_jinpin/atuo_do_daice_sw'
#window下运行的
#dirPath = os.getcwd()
print("dirPath",dirPath)

def _format_addr(s):
    name, addr = parseaddr(s)
    return formataddr((Header(name, 'utf-8').encode(), addr))
class atuo_do_daice_sw:
    _ocs_link_Prefix = _ocs_base_link + '/tv/Tasks/view/'
    _ocs_base_link = 'https://ocs.gz.cvte.cn'
    _ocs_find_old_sw_url = _ocs_base_link + "/tv/pop/selectors/common_view/Controller:FirmwareAttachments/" \
                                            "Parameters:accountId/input_id:old_fw_id/input_name:old_fw_name/inputtype:1"
    _logger = logging.getLogger(__name__)
    def __init__(self):
        self._logger.setLevel(level=logging.DEBUG)  # 控制打印级别
    def get_new_order_searchid(self,advanced_search,condition):
        #根据摘要里的条件获取search_id
        searchid = PyocsSoftware().get_searchid_from_advanced_search(advanced_search, condition)
        return searchid



    def write_ocs_have_support_auto(self,ocs_number):
        # Workbook属于创建才用
        # 打开Excel文件读取数据
        
        #用vscode运行时，用这个
        #path = dirPath+'/customers/customer_jinpin/atuo_do_daice_sw/已提交自动编译.xlsx'
        #用windows自动运行时用这个路径，并需要把pyocs放到python安装的lib\site-packages
        path = dirPath+'/已提交自动编译.xlsx'

        writebook = load_workbook(path)
        # 获取读入的文件的sheet
        sheet = writebook.active
        max_rows_new = sheet.max_row
        row_index = 1
        print("max_rows_new:",max_rows_new)
        while row_index <= max_rows_new:
            b = sheet.cell(row=row_index, column=1)
            row_index = row_index+1
            if None != b.value:
                if ocs_number == b.value:
                    return False

        row = [ocs_number]
        print(row)
        sheet.append(row)
        writebook.save(path)
        return True
    
    def write_ocs_have_submit_reuse(self,ocs_number):
        # Workbook属于创建才用
       
        #用vscode运行时，用这个
        #path = dirPath+'/customers/customer_jinpin/atuo_do_daice_sw/已提交引用.xlsx'
        #用windows自动运行时用这个路径，并需要把pyocs放到python安装的lib\site-packages
        path = dirPath+'/已提交引用.xlsx'
        writebook = load_workbook(path)
        # 获取读入的文件的sheet
        sheet = writebook.active
        max_rows_new = sheet.max_row
        row_index = 1
        print("max_rows_new:",max_rows_new)
        while row_index <= max_rows_new:
            b = sheet.cell(row=row_index, column=1)
            row_index = row_index+1
            if None != b.value:
                if ocs_number == b.value:
                    return False

        row = [ocs_number]
        print(row)
        sheet.append(row)
        writebook.save(path)
        return True

    def change_list_to_2D_list(self,new_ocs_list):
        new_ocs_order_list = list()
        i = 0
        parse_list_width = 6 
        for j in range(0,len(new_ocs_list),parse_list_width):
            new_ocs_order_list.append(new_ocs_list[j:j+parse_list_width])
            i=i+1
        return i,new_ocs_order_list

    def parse_string_from_ocs_info(self,ocs_subjest_info,string_pattern):
        pattern = re.compile(string_pattern, re.S)
        print(pattern)
        items = re.findall(pattern, ocs_subjest_info)
        print(items)
        return items

    def submit_auto_compile(self,ocs_number_list,nums):
        url = "https://tvgateway.gz.cvte.cn/bff-tv-jenkins/api/v1.0/build/build-job"


        user_info = pyocs_login.PyocsLogin.get_account_from_json_file()
        tokens = json.loads(json.dumps(eval(user_info["jenkins_token"])))



        
        check_if_have_submit_auto_flag = False
        print("submit_auto_compile ocs_number_list",ocs_number_list)

 
        for i in range(0,nums):
            ocs_number = ocs_number_list[i][0]
            if "TCL" not in ocs_number_list[i][3]:# and "[虚拟]" not in ocs_number_list[i][3]:
                auto_make_bin_flag = do_daice_task.check_if_auto_make_bin(ocs_number_list[i][4])
                surport_auto_compile_flag = do_daice_task.check_if_project_support_auto_compile(ocs_number_list[i][4])
                check_if_have_submit_auto_flag =self.write_ocs_have_support_auto(ocs_number)
                #if (check_if_have_submit_auto_flag is True) and (surport_auto_compile_flag is True):
                #surport_auto_compile_flag获取的是string类型
                if (check_if_have_submit_auto_flag is True) and (str(surport_auto_compile_flag) == "True"):
                #if  (str(surport_auto_compile_flag) == "True"):
                    print("submit_auto_compile",ocs_number)
                    jenkins_server = jenkins.Jenkins
                    if("9632" in ocs_number_list[i][4] ) or ("9256" in ocs_number_list[i][4] ) or ("9255" in ocs_number_list[i][4] ):
                        authorization = user_info["Username"] + ":" + tokens["tvci3"]

                        headers = {
                                'Content-Type': 'application/json',
                                'Authorization': base64.b64encode(authorization.encode())
                        }
                        payload = json.dumps({
                                "masterName": "tvci3",
                                "jobUri": "yangpingbu/",
                                "jobName": "Sample_Compile",
                                "params": {
                                        "OCS_NUMBER": int(ocs_number),
                                        "AUTO_MAKE_BIN": str(auto_make_bin_flag),
                                        "COMPILE_SOFTWARE": "True"
                                }
                        })
                      
                    else:
                        authorization = user_info["Username"] + ":" + tokens["tvci"]

                        headers = {
                                'Content-Type': 'application/json',
                                'Authorization': base64.b64encode(authorization.encode())
                        }                        
                        payload = json.dumps({
                                "masterName": "tvci",
                                "jobUri": "yangpingbu/",
                                "jobName": "Sample_Compile",
                                "params": {
                                        "OCS_NUMBER": int(ocs_number),
                                        "AUTO_MAKE_BIN": str(auto_make_bin_flag),
                                        "COMPILE_SOFTWARE": "True"
                                }
                        })
                    ret = requests.request("POST", url, headers=headers, data=payload)  

    def submit_atuo_reuse(self,src_ocs_number,dst_ocs_number):
        jenkins_server = jenkins.Jenkins
        account = pyocs_login.PyocsLogin().get_account_from_json_file()
        jenkins_server = jenkins.Jenkins('https://tvci.gz.cvte.cn/', username=account['Username'], password=account['Password'],
                                    timeout=30)
        print(jenkins_server)
        data = {
            'SRC_OCS_NUMBER': int(src_ocs_number),
            'DST_OCS_NUMBER': int(dst_ocs_number),
        }
        #print(data)
        r_jenkins=jenkins_server.build_job("Utility_Tools/reuse_stock_software" ,data)



    def send_email_for_differ_liaohao(self,new_ocs_info_list,count_nums):
        #从待录入需求订单表中。获取评审单号和料号。
        #以评审单号作为搜索条件获取相同评审单号和料号的订单列表。待录入需求订单的料号和其他的是否一致？有不一样的邮件提醒工程师或者交管。
        #ocs_request = pyocs_demand.PyocsDemand(ocs)
        #sw_engineer = ocs_request.get_ocs_software_engineer()
        #sw_lock_receivers = ps.get_email_addr_from_ocs(sw_engineer)
        #content = "以下订单全部禁用，请及时处理：\n" + self._ocs_id_base_link + ocs
        #mail.send_email(self.sw_lock_sender, sw_lock_receivers, self.sw_lock_cc, self.sw_lock_subject, content)
        #如果没有单要处理，old_ocs_number没有赋值会导致最后return错误
        old_ocs_number = 0
        dst_ocs_number = 0
        sw = pyocs_software.PyocsSoftware()
        for i in range(0,count_nums):

            #每个新的单重新赋值，归零可以引用的旧软件OCS,以免用了旧的值
            old_ocs_number = 0
            #self._logger.info("ocs_number not found",old_ocs_number)
            #现在是不重复reuse.但是后面做获取禁用软件信息申请解禁。还要可以继续reuse
            string_pattern_liaohao='\u5ba2\u6599\u53f7(.*?)，'
            string_pattern_customer_order='\u5ba2\u6279\u53f7(.*?)）'
            ocs_subjest_info  = u"[B.NEW][金品]TP.ATM30.PB818C配ST3751A07-2（中国，SY202005150001，客料号R41-920CN230C0AWA1C，客批号TCL海外电子01-202005-101）"
            ocs_subjest_info  = new_ocs_info_list[i][3]
            #待录入需求的虚拟单绝对可以跳过这个功能。备料和另购的因为识别的客批号不标准，这些也是客户自己消化的理论上也可以跳过。
            if ('[虚拟]' in ocs_subjest_info) or('另购' in ocs_subjest_info) or ('备料' in ocs_subjest_info):
                continue
            string_liaohao=do_daice_task.parse_string_from_ocs_info(ocs_subjest_info,string_pattern_liaohao)
            #一开始string_liaohao是list类型，要剃干净，没有一点多余的字符才行
            string_liaohao =tr(string_liaohao)
            string_liaohao = string_liaohao[2:-2]
            print("get liaohao",string_liaohao)
            #客批号
            #另购、备料，不规范，为空的情况多，直接去掉
            string_customer_order=do_daice_task.parse_string_from_ocs_info(ocs_subjest_info,string_pattern_customer_order)
            #一开始string_liaohao是list类型，要剃干净，没有一点多余的字符才行
            string_customer_order =str(string_customer_order)
            string_customer_order = string_customer_order[2:-2]            
            customer_order_advanced_search = {
                "0":{"search_field_name":"Task.subject","search_field_id":"554","search_field_type":"5","search_field_rel_obj":"null","search_opr":"TDD_OPER_INC","input1":string_customer_order,"input2":"null","offset":"null"}
                }
            new_order_condition = ""
            search_id_customer_order = do_daice_task.get_new_order_searchid(customer_order_advanced_search,new_order_condition)
            print("search_id_customer_order",search_id_customer_order)
            xpath_result_customer_order = PyocsList().get_single_data_list(list_xpath,search_id_customer_order)
            temp_list=do_daice_task.change_list_to_2D_list(xpath_result_customer_order[1])
            ocs_list_customer_order = temp_list[1]
            count_nums = temp_list[0]
            ocs_list= temp_list[1]
            print(temp_list)
            print("count_nums",count_nums)
            if count_nums >1 :
            ##if(new_ocs_info_list[i][4] != "MT7601") and (new_ocs_info_list[i][4] != "6M60") and ("MT9602" not in new_ocs_info_list[i][4]) :
                for j in range(count_nums-1,-1,-1):
                    #之前待录入需求的订单料号，不在搜索到的订单列表里面，说明有相同订单号的料号不一样
                    
                    if string_liaohao not in ocs_list[j][3]:
                        dst_ocs_number = str(ocs_list[j][0])
                        print("不同客料号的订单",dst_ocs_number)
                        #break
                    #去掉相同评审单号的虚拟单
                    if dst_ocs_number !=0 and '虚拟' not in ocs_list[j][3]:
                    ##if dst_ocs_number !=0 :
                        ocs_request = PyocsDemand(dst_ocs_number)
                        sw_engineer = ocs_request.get_ocs_software_engineer()
                        ##receivers = sw.get_email_addr_from_ocs(sw_engineer)
                        ##print("receivers",receivers)
                        order_neiqing = ocs_request.get_nq()
                        receivers = sw.get_email_addr_from_ocs(order_neiqing)
                        ##cc_email  = "linxiangna156@163.com"
                        subject_str = '相同评审单号有不同的料号提醒'                
                        content = "该评审单号发现相同评审单号有不同的料号，会造成软件处理异常，请帮忙及时处理：\n" + string_customer_order

                        msg = MIMEText(content, 'plain', 'utf-8')  # text_str发送正文，固定语句+确认不通过的订单信息                        
                        msg['From'] = _format_addr('CVTE_林祥纳<%s>' % from_addr)
                        #msg['To'] = ';'.join(receivers)  # 以逗号形式连接起来 #_format_addr('管理员 <%s>' % to_addr)#可以采用把列表分开来的方法，一个一个发
                        msg['To'] = Header(receivers) 
                        ##msg['Cc'] = Header(cc_email)
                        msg['Subject'] = Header(subject_str, 'utf-8').encode()

                        server = smtplib.SMTP(smtp_server, 25)
                        server.set_debuglevel(1)
                        server.login(from_addr, password)
                        server.sendmail(from_addr, receivers, msg.as_string())
                        server.quit()
                        

    def check_if_auto_make_bin(self,str_fangan):
        
        #window下运行时，用这个
        _support_auto_make_bin_file = dirPath+'/support_auto_make_bin.json'
        try:
            logging.info(_support_auto_make_bin_file)
            with open(_support_auto_make_bin_file, 'r') as load_file:
                tmp = load_file.read()
                support_auto_make_bin_dict = json.loads(json.dumps(eval(tmp)))
                if  str(str_fangan) in support_auto_make_bin_dict:
                    return support_auto_make_bin_dict[str_fangan]
                else:
                    return False                   
                #return support_auto_make_bin_dict
        except FileNotFoundError:
            logging.error("请查找运行路径是否有保存支持自动做bin方案的json文件")
            #先默认支持做bin.支持做bin没什么坏处
            return True

    def check_if_project_support_auto_compile(self,str_fangan):
        
        #window下运行时，用这个
        #getcwd()获取行时所在目录，对运行路径要求苛刻
        _support_auto_make_bin_file = dirPath+'/support_auto_compile.json'
        try:
            logging.info(_support_auto_make_bin_file)
            with open(_support_auto_make_bin_file, 'r') as load_file:
                tmp = load_file.read()
                support_auto_compile_dict = json.loads(json.dumps(eval(tmp)))
                if str(str_fangan) in support_auto_compile_dict:
                    return support_auto_compile_dict[str_fangan]
                else:
                    return False
                #return support_auto_make_bin_dict
        except FileNotFoundError:
            logging.error("请查找运行路径是否有保存支持自动编译方案的json文件")
            #先默认支持自动编译.支持自动编译稍微浪费了点服务器资源
            return True

    def parse_lock_software_and_reason(self,ocs_list):
        flag = True
        return
    #还需要加功能，判断订单返回的可用软件列表不为空才应用
    #同一料号还有占空比不一样的情况（较少）
    #找旧软件应该要尽量找已经测试通过的
    def check_if_reuse_old(self,new_ocs_info_list,count_nums):
        #如果没有单要处理，old_ocs_number没有赋值会导致最后return错误
        old_ocs_number = 0
        sw = pyocs_software.PyocsSoftware()
        for i in range(0,count_nums):
            dst_ocs_number = str(new_ocs_info_list[i][0])
            print("dst_ocs_number",dst_ocs_number)
            pd_dst = PyocsDemand(ocs_number=dst_ocs_number)
            print("产品型号",pd_dst.get_product_name())
            dst_product_name = pd_dst.get_product_name()
            dst_pwm = pd_dst.get_pwm()
            dst_port_info = pd_dst.get_port_info()
            #每个新的单重新赋值，归零可以引用的旧软件OCS,以免用了旧的值
            old_ocs_number = 0
            #self._logger.info("ocs_number not found",old_ocs_number)
            #现在是不重复reuse.但是后面做获取禁用软件信息申请解禁。还要可以继续reuse
            string_pattern_liaohao='\u5ba2\u6599\u53f7(.*?)，'
            string_pattern_model_size='E(\d\d)'
            ocs_subjest_info  = u"[B.NEW][金品]TP.ATM30.PB818C配ST3751A07-2（中国，SY202005150001，客料号R41-920CN230C0AWA1C，客批号TCL海外电子01-202005-101）"
            ocs_subjest_info  = new_ocs_info_list[i][3]
            if '[虚拟]' in ocs_subjest_info:
                continue
            string_model_size  = do_daice_task.parse_string_from_ocs_info(new_ocs_info_list[i][5],string_pattern_model_size)
            string_liaohao=do_daice_task.parse_string_from_ocs_info(ocs_subjest_info,string_pattern_liaohao)
            #一开始string_liaohao是list类型，要剃干净，没有一点多余的字符才行
            string_liaohao =str(string_liaohao)
            string_liaohao = string_liaohao[2:-2].split("-", 2)[1]
            print("after split",string_liaohao)
            liaohao_advanced_search = {
                "0":{"search_field_name":"Task.subject","search_field_id":"554","search_field_type":"5","search_field_rel_obj":"null","search_opr":"TDD_OPER_INC","input1":string_liaohao,"input2":"null","offset":"null"}
                }
            new_order_condition = ""
            search_id_liaohao = do_daice_task.get_new_order_searchid(liaohao_advanced_search,new_order_condition)
            #print(search_id)
            xpath_result_liaohao = PyocsList().get_single_data_list(list_xpath,search_id_liaohao)
            temp_list=do_daice_task.change_list_to_2D_list(xpath_result_liaohao[1])
            ocs_list_liaohao = temp_list[1]
            count_nums = temp_list[0]
            #print(temp_list)
            print("count_nums",count_nums)
            if(new_ocs_info_list[i][4] != "MT7601") and (new_ocs_info_list[i][4] != "6M60") and ("MT9602" not in new_ocs_info_list[i][4]) :
                for j in range(count_nums-1,-1,-1):
                    print("第几轮引用j:",j)
                    #现在是找ocs上有未被禁用软件的订单
                    print("ocs_list_liaohao[j][2]",ocs_list_liaohao[j][2])
                    if ocs_list_liaohao[j][2] != "软件审核不通过" and ocs_list_liaohao[j][2] != "软件测试不通过" and ocs_list_liaohao[j][2] != "待录入需求":
                        #if do_daice_task.get_enable_software_can_be_reuse_with_ocs(ocs_list_liaohao[j][0]) == None:
                        if sw.get_enable_software_list(ocs_list_liaohao[j][0]) == None:                        
                            print("get_enable_software_can_be_reuse_with_ocs NONE")
                            continue
                        else:
                            print("ocs_list_liaohao[j][0]",ocs_list_liaohao[j][0])
                            #print("get_enable_software_reuse_ocs_with_ocs",PyocsSoftware().get_enable_software_reuse_ocs_with_ocs(ocs_list_liaohao[j][0]))
                            old_ocs_string_model_size  = do_daice_task.parse_string_from_ocs_info(ocs_list_liaohao[j][5],string_pattern_model_size)
                            #print(old_ocs_string_model_size)
                            old_backup_ocs_number = str(ocs_list_liaohao[j][0])
                            pd_old = PyocsDemand(ocs_number=old_backup_ocs_number)
                            print("库存单的产品型号",pd_old.get_product_name())
                            old_product_name = pd_old.get_product_name()
                            old_pwm = pd_old.get_pwm()
                            old_port_info = pd_old.get_port_info()
                            #if  old_ocs_string_model_size == string_model_size:
                            if  (dst_product_name == old_product_name) and (dst_pwm == old_pwm) and (dst_port_info == old_port_info):
                                old_ocs_number= ocs_list_liaohao[j][0]
                                break
            print("old_ocs_number",old_ocs_number)
            if old_ocs_number!=0:
                #do_daice_task.submit_atuo_reuse(old_ocs_number,dst_ocs_number)
                workspace = os.getcwd()
                #这个reuse函数出问题了？
                ret=PyocsSoftware().reuse_old_sw_from_src_to_dst(old_ocs_number,dst_ocs_number,workspace)
        return old_ocs_number

    def get_enable_software_can_be_reuse_with_ocs(self, ocs, exclude_bin=True):
        """获取订单上的启用状态下的软件以及对应的信息字典
        Args:
            html: 某个ocs订单的html页面
        Returns:
            启用状态下的非烧录 bin 或者不符合 CVTE 规范的软件以及对应的信息字典
        """
        active_software_info_list = []
        _ocs_base_link = 'https://ocs.gz.cvte.cn'
        _ocs_link_Prefix = _ocs_base_link + '/tv/Tasks/view/'
        src_ocs_link = _ocs_link_Prefix + str(ocs)
        response = PyocsRequest().pyocs_request_get(src_ocs_link)
        html = etree.HTML(response.text)
        enable_software_name_xpath = \
            '//a[contains(@href, "/tv/Attachments/download_attachment")]/text()'
        enable_software_name_list = html.xpath(enable_software_name_xpath)
        print("ocs,enable_software_name_list",ocs,enable_software_name_list)
        for sw_name in enable_software_name_list:
            if not PyocsSoftware().is_sw_locked(sw_name=sw_name, html=html):
                active_software_info_list.append(sw_name)
        print("active_software_info_list",ocs,active_software_info_list)
        if active_software_info_list:
            enable_reuse_software_info = active_software_info_list[0]
            enable_reuse_ocs_list = re.findall(r'\d+', enable_reuse_software_info)
            return active_software_info_list
        else:
            return None

    def get_en_sw_list_from_html(self, html):
        """获取订单上的启用状态下的软件列表，包含被锁定的软件
        Args:
            html: 某个ocs订单的html页面
        Returns:
            启用状态下的软件列表
        """
        enable_software_list = []
        enable_software_name_xpath = \
            '//a[contains(@href, "/tv/Attachments/download_attachment")]/text()'
        enable_software_name_list = html.xpath(enable_software_name_xpath)
        print("启用状态的软件:" ,enable_software_name_list)
        for sw_name in enable_software_name_list:
            enable_software_list.append(sw_name)
        if enable_software_list:
            print("enable_software_list:" + str(enable_software_list))            
            return enable_software_list
        else:
            return None

class Logger(object):
    def __init__(self, filename='jinpindaice.log', stream=sys.stdout):
        self.terminal = stream
        self.log = open(filename, 'a')

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        pass

sys.stdout = Logger(stream=sys.stdout)
sys.stdout = Logger(stream=sys.stderr)

sys_time=time.strftime('time:%Y-%m-%d-%H-%M-%S', time.localtime(time.time()))
print(sys_time)
do_daice_task = atuo_do_daice_sw()
pyocs_job = PyocsSoftware()

#金品 且待录入需求的订单
new_order_advanced_search = {
    "0":{"search_field_name":"Task.account_id","search_field_id":"560","search_field_type":"19","search_field_rel_obj":"Accounts","search_opr":"TDD_OPER_INC","input1":"\u91d1\u54c1","input2":"null","offset":"null"},
    "1":{"search_field_name":"Task.status","search_field_id":"584","search_field_type":"19","search_field_rel_obj":"Enums","search_opr":"TDD_OPER_INC","input1":ocs_status,"input2":"null","offset":"null"}
    }
#按成员来
#new_order_advanced_search = {
    #"0": {"search_field_name":"Task.sw_user_id","search_field_id":"555","search_field_type":"19","search_field_rel_obj":"Users","search_opr":"TDD_OPER_INC","input1":member,"input2":"null","offset":"null"},
    #"1": {"search_field_name":"Task.status","search_field_id":"584","search_field_type":"19","search_field_rel_obj":"Enums","search_opr":"TDD_OPER_EQUAL","input1":ocs_status,"input2":"null","offset":"null"}
    #}

new_order_condition = "(1 and 2)"

#print(do_daice_task.check_if_project_support_auto_compile("MS3683"))

logging.basicConfig(filename=os.path.join(os.getcwd()+'/customers/customer_jinpin/atuo_do_daice_sw/','log.txt'),level=logging.DEBUG)
logging.debug('this is a message')
search_id_new_order = do_daice_task.get_new_order_searchid(new_order_advanced_search,new_order_condition)
print(search_id_new_order)

xpath_result = PyocsList().get_single_data_list(list_xpath,search_id_new_order)
#print(xpath_result)
new_ocs_list = xpath_result[1]
print(new_ocs_list)
#count_ocs_numbers = PyocsList().get_ocs_id_list(search_id)
temp_list=do_daice_task.change_list_to_2D_list(new_ocs_list)
new_ocs_info_list = temp_list[1]
count_nums = temp_list[0]
#print(count_nums)
#print(new_ocs_info_list[0])
"""
_ocs_base_link = 'https://ocs.gz.cvte.cn'
_ocs_link_Prefix = _ocs_base_link + '/tv/Tasks/view/'
src_ocs_link = _ocs_link_Prefix + str(690639)
response = PyocsRequest().pyocs_request_get(src_ocs_link)
html = etree.HTML(response.text)
do_daice_task.get_en_sw_list_from_html(html) 
#do_daice_task.get_enable_software_can_be_reuse_with_ocs("690639") 
"""

#do_daice_task.send_email_for_differ_liaohao(new_ocs_info_list,count_nums)

do_daice_task.check_if_reuse_old(new_ocs_info_list,count_nums)



#reuse成功后，应该就没有待录入需求目录了
#大致等个10分钟，等reuse完。改方法不是很好。应该是直接获取reuse成功状态
#time.sleep(600)
search_id_new_order = do_daice_task.get_new_order_searchid(new_order_advanced_search,new_order_condition)
print(search_id_new_order)

xpath_result = PyocsList().get_single_data_list(list_xpath,search_id_new_order)
#print(xpath_result)
new_ocs_list = xpath_result[1]
print(new_ocs_list)
#count_ocs_numbers = PyocsList().get_ocs_id_list(search_id)
temp_list=do_daice_task.change_list_to_2D_list(new_ocs_list)
new_ocs_info_list = temp_list[1]
count_nums = temp_list[0]
do_daice_task.submit_auto_compile(new_ocs_info_list,count_nums)



