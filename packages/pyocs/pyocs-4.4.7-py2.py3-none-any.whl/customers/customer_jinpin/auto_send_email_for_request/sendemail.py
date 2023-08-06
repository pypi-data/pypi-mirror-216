#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import smtplib
from email import encoders
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import parseaddr, formataddr
import sys
sys.path.append(".")
from openpyxl import Workbook
from openpyxl import load_workbook
from pyocs import pyocs_login


def _format_addr(s):
    name, addr = parseaddr(s)
    return formataddr((Header(name, 'utf-8').encode(), addr))



# Workbook属于创建才用
# 打开Excel文件读取数据
dirPath = os.getcwd()
readbook = load_workbook(dirPath+'/customers/customer_jinpin/auto_send_email_for_request/追需求表格2.xlsx')
#account = pyocs_login.PyocsLogin().get_account_from_json_file()
#from_addr = str(account['Username'])+"@cvte.com"
#password = account['Password']
from_addr = 'linxiangna@cvte.com'
password = 'Lin13690439782'
to_engyneer = ['linxiangna@cvte.com',
               'xiaozhitao@cvte.com', 'leimingsheng@cvte.com','wangjinxin@cvte.com',
                'yunfeng.chang@jpe.cc', 'xiaocong.cen@jpe.cc','wang.chen@jpe.cc','guozhao.guan@jpe.cc',
               'xiaoping.liao@jpe.cc', 'kaijian.lin@jpe.cc','guixuan.zheng@jpe.cc','zhengding.ji@jpe.cc','xuebo.luo@jpe.cc']

#to_engyneer = ['linxiangna@cvte.com']
# to_engyneer=sned_to_target.split(",")
smtp_server = 'smtp.exmail.qq.com'
text_str1 = 'Dear  各位金品电子工程师：'
text_str2 = '\n'
text_str3 = '     您们好！请提供如上软件需求及参考软件版本，如果需求已提供请忽略此邮件，谢谢！'
subject_str1 = '【新订单处理 】'
subject_str2 = '祥鑫（BTY）   '
subject_str3 = '机型:E55EK2100'
subject_str4 = 'TP.HV553.PC821(4K*2K)'
subject_str5 = '评审单号:'
subject_str6 = '01-201903-105'
subject_str7 = '孟加拉'
subject_str8 = '     '
subject_str9 = '客料号'
kongge = '  '
keliaohao = '  '

str_fangan="     方案对应工程师如下："
str_chenchaoxiong="  王金鑫 516(9256)  ,518(9255)"
str_linxiangna="     林祥纳 ATM30(920),708(972),3553,9602,105(6710),105S(512C)"
str_chenfan="      雷明生 56,53(108),69,706(9632),508,506,3663,3683"
str_fangan_beizhu=text_str2+str_fangan+text_str2+str_chenchaoxiong+text_str2+str_linxiangna+text_str2+str_chenfan+text_str2
text_str = text_str1 + text_str2 + text_str3+str_fangan_beizhu
#text_str="Add 国钊"
# 打开excel
#readbook = load_workbook(dir_str1)
# 获取读入的文件的sheet
sheet = readbook.active
max_rows_new = sheet.max_row
row_index = 1
print(max_rows_new)
while row_index < max_rows_new:
    b = sheet.cell(row=row_index, column=1)

    if None != b.value:
        subject_str2 = sheet.cell(row=row_index, column=6).value
        subject_str7 = sheet.cell(row=row_index, column=7).value
        subject_str3 = sheet.cell(row=row_index, column=9).value
        subject_str4 = sheet.cell(row=row_index, column=13).value
        keliaohao   = sheet.cell(row=row_index, column=12).value
        subject_str6 = b.value
        if(subject_str7==None):
            subject_str7="国家未定"
        # 不接受有这些关键项为空的情况
        subject_str = subject_str1 + subject_str2 + kongge + subject_str7 + subject_str8 + subject_str3 +\
                      subject_str8 + subject_str4 + subject_str8 + subject_str5 + subject_str6+kongge+subject_str9+keliaohao
        #msg = MIMEMultipart('alternative')
        msg = MIMEText(text_str, 'plain', 'utf-8')  # text_str发送正文，固定语句+确认不通过的订单信息
        #msgTEXT2 = MIMEText(str_fangan_beizhu, 'plain', 'utf-8')
        #msg.attach(msgTEXT1)
        #msg.attach(msgTEXT2)
        #msg_co = MIMEText(text_str, 'plain', 'utf-8')
        msg['From'] = _format_addr('CVTE_林祥纳<%s>' % from_addr)
        msg['To'] = ';'.join(to_engyneer)  # 以逗号形式连接起来 #_format_addr('管理员 <%s>' % to_addr)#可以采用把列表分开来的方法，一个一个发
        msg['Subject'] = Header(subject_str, 'utf-8').encode()

        server = smtplib.SMTP(smtp_server, 25)
        server.set_debuglevel(1)
        server.login(from_addr, password)
        server.sendmail(from_addr, to_engyneer, msg.as_string())
        print(from_addr)
        print(to_engyneer)
    row_index += 1
server.quit()
