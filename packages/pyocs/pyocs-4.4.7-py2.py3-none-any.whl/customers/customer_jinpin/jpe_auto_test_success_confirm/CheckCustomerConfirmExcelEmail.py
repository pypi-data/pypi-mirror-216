import datetime
import email
import poplib
poplib._MAXLINE=204800#重设行长度
import email.policy
import telnetlib
import time
import os
import json
import re
import sys
from openpyxl import load_workbook
from email.parser import Parser
from email.header import decode_header
from email.utils import parseaddr
from pyocs import pyocs_login
import global_var as gl    #添加全局变量管理模块
pop3_server = 'pop.cvte.com'
poplib._MAXLINE=20480
#=============================================================
# 脚本介绍和说明：
# 1、在CVTE企业网页邮箱里设置把收取3个月的邮件改成收取全部
# 2、这个脚本第一次运行会生产一个文件counttxt.txt记录目前邮件数
# 3、下次运行会获取最新的邮件的数，然后处理新增的邮件
#==============================================================
gl._init()
class CheckCustomerConfirmExcelEmail:

    subject_title="JPE软件确认"
    account = pyocs_login.PyocsLogin().get_account_from_json_file()
    from_addr = str(account['Username']) + "@cvte.com"
    #password = account['Password']

    email_user = 'linxiangna@cvte.com'
    password ='Lin13690439782'
    dirPath = os.getcwd()

    #记录邮件个数文件的路径
    counttxtpath= dirPath+'/customers/customer_jinpin/jpe_auto_test_success_confirm/counteConfirmEmail.xlsx'
    AttachName = "回复CVT软件信息"#"软件确认汇总"
    # 在指定目录下创建文件，注意二进制文件需要用wb模式打开
    confirmfilepath=dirPath+'/customers/customer_jinpin/jpe_auto_test_success_confirm/confirm_excel/'
    save_confirm_file_name=dirPath+'/customers/customer_jinpin/jpe_auto_test_success_confirm/confirm_file_name.txt'
    
    new_confirm_excel_name=""
    save_confirm_sw_file=dirPath+'/customers/customer_jinpin/jpe_auto_test_success_confirm/'
    
    # 解析邮件,获取附件
    def get_att(self,msg_in):
        # import email
        attachment_files = []
        for part in msg_in.walk():
            # 获取附件名称类型
            file_name = part.get_filename()
            if file_name:
                h = email.header.Header(file_name)
                # 对附件名称进行解码
                dh = email.header.decode_header(h)
                filename = dh[0][0]
                if dh[0][1]:
                    # 将附件名称可读化
                    filename = self.decode_str(str(filename, dh[0][1]))
                    print("附件名:"+str(filename))
                    if self.AttachName in str(filename):  
                        # filename = filename.encode("utf-8")
                        # 下载附件
                        data = part.get_payload(decode=True)
                        print(filename)
                        att_file = open(self.confirmfilepath+filename, 'wb')
                        attachment_files.append(filename)
                        att_file.write(data)  # 保存附件
                        att_file.close()
                        fw = open(self.save_confirm_file_name, 'w',encoding='utf-8',errors='ignore')
                        fw.write(str(filename))
                        fw.close() 
                        #global recive_new_email_flag
                        gl.set_recive_new_email_flag(True)
                        print("recive",gl.get_recive_new_email_flag())                  
                        print("找到了，退出！")
                        exit()
        return attachment_files
    # 解析邮件,获取附件
    def get_str(self,msg_in):
        # import email
        attachment_files = []
        for part in msg_in.walk():
            # 获取附件名称类型
            file_name = part.get_filename()
            if file_name:
                h = email.header.Header(file_name)
                # 对附件名称进行解码
                dh = email.header.decode_header(h)
                filename = dh[0][0]
                if dh[0][1]:
                    # 将附件名称可读化
                    filename = self.decode_str(str(filename, dh[0][1]))
                    print("附件名:"+str(filename))
                    if self.AttachName in str(filename):  
                        # filename = filename.encode("utf-8")
                        # 下载附件
                        data = part.get_payload(decode=True)
                        print(filename)
                        att_file = open(self.confirmfilepath+filename, 'wb')
                        attachment_files.append(filename)
                        att_file.write(data)  # 保存附件
                        att_file.close()
                        fw = open(self.save_confirm_file_name, 'w',encoding='utf-8',errors='ignore')
                        fw.write(str(filename))
                        fw.close() 
                        #global recive_new_email_flag
                        gl.set_recive_new_email_flag(True)
                        print("recive",gl.get_recive_new_email_flag())                  
                        print("找到了，退出！")
                        exit()
        return attachment_files
    #==============================================================
    # 解析邮件必要的一些字符转换
    #==============================================================
    def decode_str(self, s):
        value, charset = decode_header(s)[0]
        if charset:
            try:
                value = value.decode("gb2312")
            except:
                return
            #value = value.decode(charset)
        return value

    def guess_charset(self, msg):
        charset = msg.get_charset()
        if charset is None:
            content_type = msg.get('Content-Type', '').lower()
            pos = content_type.find('charset=')
            if pos >= 0:
                charset = content_type[pos + 8:].strip()
        return charset

    def get_mail_text(self, msg, indent=0):
        # indent用于缩进显示
        text = ''
        if indent == 0:
            Che = CheckCustomerConfirmExcelEmail()
            try:
                subject_str = Che.decode_str(msg.get('Subject',''))
            except UnicodeDecodeError:
                print("exception test error")
                return
            else:
                if Che.subject_title in str(subject_str):
                    for header in ['From', 'To', 'Subject']:
                        value = msg.get(header, '')
                        print("value",value)
                        if value:
                            if header=='Subject':
                                value = self.decode_str(value)
                            else:
                                hdr, addr = parseaddr(value)
                                name = self.decode_str(hdr)
                                value = u'%s <%s>' % (name, addr)

                        text += '%s%s: %s\n' % ('  ' * indent, header, value)
                else:#非确认邮件，调过获取
                    text += '%s%s: %s\n' % ('  ' * indent, '', '')
        if (msg.is_multipart()):
            parts = msg.get_payload()
            for n, part in enumerate(parts):
                text += self.get_mail_text(part, indent + 1)
        else:
            content_type = msg.get_content_type()
            if content_type == 'text/plain'  or content_type=='text/html':
                content = msg.get_payload(decode=True)
                charset = self.guess_charset(msg)
                if charset:
                    try:
                        content = content.decode("gb2312")
                    except:
                        text = ""
                        return text
                    #content = content.decode(charset)
                text += '%sText: %s' % ('  ' * indent, content)
        return text

    def get_base_sw_name_and_config_name(self, text):
        #regular = re.compile(r'基础软件名：([\S]*)')    
        #tmp = re.findall(regular, text)[0]
        #filter_text = re.findall(regular, text)[0]      
        #print("filter_text",filter_text)        
        #for jinpin , get the base sw name and config name from mail
        regular2 = re.compile(r'\d{8}_\d{6}')
        #直接根据这个规则find ，会find到6个数值
        all_base_name = re.findall(regular2,text)
        print("all_base_name",all_base_name)
        #金品foxmail的识别会为空，不会往下跑了，金品的邮件真的有配置包的情况会继续往下跑，自己发的邮件没有配置包也会识别出两个
        if len(all_base_name) > 1:
            print("all_base_name[1]",all_base_name[1])
            #base_name = re.findall(regular, text)[0][0:15]
            base_name = re.findall(regular2, text)[0]

            print("base_name",base_name)
            #判断获取到的软件版本，第一个和第二个是不是一样，不一样则是有两个
            if str(base_name) != str(all_base_name[1]):
                return [base_name, all_base_name[1]]
            else:
                return [base_name, '']
        else:
            return [all_base_name, '']                

    def get_base_sw_name(self, text):
        #金品专用，拿到评审单号
        #regular = re.compile(r'(\[CVTE备货.*三部.*?)（')
        #regular = re.compile(r'TWBZ[\w|-]*\.\w*\.')#订单号有TWBC的情况
        #匹配任意非空白字符，现在看起来是可行的，格式变化了可能就又出问题了
        regular = re.compile(r'基础软件名：([\S]*)')    
        #tmp = re.findall(regular, text)[0]
        tmp = re.findall(regular, text)
        print("get_base_sw_name tmp =",tmp)
        return tmp
        if tmp:
            return tmp[0]
        else:
            return ''        

    def get_config_sw_name(self, text):
        #金品专用，拿到评审单号
        #regular = re.compile(r'(\[CVTE备货.*三部.*?)（')
        #regular = re.compile(r'TWBZ[\w|-]*\.\w*\.')#订单号有TWBC的情况
        #匹配任意非空白字符，现在看起来是可行的，格式变化了可能就又出问题了
        regular = re.compile(r'包软件：([\S]*)')    
        #tmp = re.findall(regular, text)[0]
        tmp = re.findall(regular, text)
        print("get_config_sw_name tmp =",tmp)
        if tmp:
            return tmp[0]
        else:
            return ''

    def get_just_sw_number(self, text):
        #regular = re.compile(r'基础软件名：([\S]*)')    
        #tmp = re.findall(regular, text)[0]
        #filter_text = re.findall(regular, text)[0]      
        #print("filter_text",filter_text)        
        #for jinpin , get the base sw name and config name from mail
        regular = re.compile(r'\d{8}_\d{6}')
        tmp = re.findall(regular, text)
        print("get_just_sw_number tmp =",tmp)
        if tmp:
            return tmp[0]
        else:
            return ''

    def get_vidaa_version(self, text):

        #regular = re.compile(r'）(*.?)<')
        #tmp = re.findall(regular, text)
        tmp = text.split('<br')[0].rsplit('）')[0].rsplit(')')[0]
        print("get_just_sw_number tmp =",tmp)
        if tmp:
            return tmp[0]
        else:
            return text             

    def old_get_base_sw_name_and_config_name(self, text):
        #for jinpin , get the base sw name and config name from mail
        regular = re.compile(r'\d{8}_\d{6}')
        base_name = re.findall(regular, text)[0][0:15]
        if len(base_name) > 1:
            return [base_name, base_name[1]]
        else:
            return [base_name, '']

    def get_oder_num_and_board(self, text):
        #金品专用，拿到评审单号，主板型号
        regular = re.compile(r'TWBZ[\w|-]*\.\w*\.')#订单号有TWBC的情况
        #tmp = re.findall(regular, text)[0]
        tmp = re.findall(regular, text)
        print("tmp =",tmp)
        return [tmp[17:30], tmp[33:38]]

    def get_board_type(self, text):
        #金品专用，拿到评审单号
        #regular = re.compile(r'(\[CVTE备货.*三部.*?)（')
        #regular = re.compile(r'TWBZ[\w|-]*\.\w*\.')#订单号有TWBC的情况
        regular = re.compile(r'主板型号：([\w|.]*)')    
        #tmp = re.findall(regular, text)[0]
        tmp = re.findall(regular, text)[0]
        print("get_board_type tmp =",tmp)
        return tmp

    def get_oder_num(self, text):
        #金品专用，拿到评审单号
        #regular = re.compile(r'(\[CVTE备货.*三部.*?)（')
        #regular = re.compile(r'TWBZ[\w|-]*\.\w*\.')#订单号有TWBC的情况
        regular = re.compile(r'评审单号：([\w|-]*)')    
        tmp = re.findall(regular, text)[0]
        #tmp = re.findall(regular, text)
        print("get_oder_num tmp =",tmp)
        return tmp

    def add_info_to_nut_excel_from_email(self, ocs, workspace):
        """
        根据待上传OCS上的客料号，从BOE出海信软件确认表中找到已确认的软件包名，并上传。
        """
        pd = PyocsDemand(ocs_number=ocs)
        customer_info = pd.get_ocs_customer()
        passenger_number = pd.get_passenger_number()
        passenger_number = passenger_number.strip()
        print("OCS订单上的客料号：" + passenger_number)

        kb = pyocs_confluence.PyocsConfluence()
        download_link = kb.get_download_link_software_confirm_table_by_customer(customer_info)
        excel_file_location = PyocsFileSystem.get_file_from_nut_driver(url=download_link, workspace=workspace)

        #从Excel表中由OCS上的客料号获取到客户确认的软件名称
        workbook = xlrd.open_workbook(filename=excel_file_location)
        table = workbook.sheet_by_index(0)
        board_type_list = table.col_values(0) #料号在确认表中的第0列
        sw_name_column = 1 #客户确认的软件名在表中的第1列
        customer_confirm_sw_name = None
        for index in range(len(board_type_list)):
            board_type_name = board_type_list[index].strip()
            if passenger_number in board_type_name:
                customer_confirm_sw_name = table.cell_value(index, sw_name_column)
                break

        #用客户确认的软件包名在OCS中查找已使用的OCS，并上传到目标OCS上
        if customer_confirm_sw_name:
            customer_confirm_sw_name = customer_confirm_sw_name.strip()
            print("客户已确认的软件包: " + customer_confirm_sw_name)
            search_form_data = dict()
            search_form_data['_method'] = 'POST'
            search_form_data['AttachmentName'] = customer_confirm_sw_name
            search_response = PyocsRequest().pyocs_request_post(self._search_url, data=search_form_data)
            html = etree.HTML(search_response.text)
            sw_used_ocs_list = html.xpath('//td/strong/text()')
            if sw_used_ocs_list != [] :
                sw_used_ocs = sw_used_ocs_list[0].strip("#")
                self.reuse_old_sw_from_src_to_dst(sw_used_ocs, ocs, workspace)
        else:
            raise NoConfirmSoftwareError("确认表中没有找到客户确认的软件！")

class Logger(object):
    def __init__(self, filename='getexcel.log', stream=sys.stdout,errors='ignore'):
        self.terminal = stream
        self.log = open(filename, 'a')

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        pass
#global recive_new_email_flag    
#Check_in_confirm = CheckCustomerConfirmExcelEmail.CheckCustomerConfirmExcelEmail()
#CheckCustomerConfirmExcelEmail Check_in_confirm = Che



#==============================================================
# 主程序
# 登录邮箱。获取邮件个数，对比处理新增邮件个数
# 将名称含有subject_title 特殊字符的邮件才进行解析处理。
#==============================================================
if __name__ == '__main__':
    # 连接到POP3服务器,有些邮箱服务器需要ssl加密，可以使用poplib.POP3_SSL
    try:
  
        sys.stdout = Logger(stream=sys.stdout)
        sys.stdout = Logger(stream=sys.stderr)
        sys_time=time.strftime('time:%Y-%m-%d-%H-%M-%S', time.localtime(time.time()))
        print(sys_time)

        telnetlib.Telnet('pop.cvte.com', 995)
        server = poplib.POP3_SSL(pop3_server, 995, timeout=100)
    except:
        time.sleep(5)
        server = poplib.POP3(pop3_server, 110, timeout=100)

    # 可以打开或关闭调试信息
    # server.set_debuglevel(1)
    # 打印POP3服务器的欢迎文字:
    print(server.getwelcome().decode('utf-8'))
    Che = CheckCustomerConfirmExcelEmail()
    #为啥没有新邮件，连身份认证都不过
    # 身份认证:
    server.user(Che.email_user)
    server.pass_(Che.password)
    # 返回邮件数量和占用空间
    print('Messages: %s. Size: %s' % server.stat())
    # list()返回所有邮件的编号:

    resp, mails, octets = server.list()

    new_email_num = len(mails)
    print("当前邮件个数 ="+str(new_email_num))

    if os.path.exists(Che.counttxtpath):
        #记录文件存在
        fr =open(Che.counttxtpath)
        old_email_num = fr.read()

        if (str(old_email_num) == str(new_email_num)):
            #没有新邮件
            print("没有新邮件")
            fr.close()
        else:
            #有新邮件，得到邮件个数差值
            dif_value = int(new_email_num) - int(old_email_num)
            fr.close()
            #记录新的邮件个数
            fw = open(Che.counttxtpath, 'w',encoding='utf-8',errors='ignore')
            fw.write(str(new_email_num))
            fw.close()

            execl_path = Che.save_confirm_sw_file+'客户回复确认表格.xlsx'
            #fr_confirm =open(Che.save_confirm_sw_file+'客户回复确认表格.xlsx','r+')
            #fr_confirm.truncate()
            Workbook = load_workbook(execl_path)
            sheet = Workbook.get_sheet_by_name('test')
            max_rows_new = sheet.max_row
            print("max_rows_new",max_rows_new)
            row_index = 1
            while row_index <= max_rows_new:
                sheet.cell(row=row_index, column=2).value = ""
                sheet.cell(row=row_index, column=3).value = ""
                sheet.cell(row=row_index, column=4).value = ""
                sheet.cell(row=row_index, column=5).value = ""
                row_index = row_index + 1
            Workbook.save(execl_path)
            #Workbook = load_workbook(execl_path)
            #sheet = Workbook.get_sheet_by_name('test')             
            confirm_line = 0
            sum_confirm_email_once_time = 0
            for index in range(int(old_email_num)+1, int(new_email_num)+1):

                try:
                    resp, mails, octets = server.list()
                    # 可以查看返回的列表类似[b'1 82923', b'2 2184', ...]
                    # print(mails)

                    # 获取最新一封邮件, 注意索引号从1开始:

                    resp, lines, octets = server.retr(index)
                except:
                    #line too long的肯定不是确认邮件
                    index= index+1
                    continue
                # lines存储了邮件的原始文本的每一行,
                # 可以获得整个邮件的原始文本:
                msg_content = b'\r\n'.join(lines).decode('utf-8','ignore')
                #msg_content = b'\r\n'.join(lines).decode('gbk')
                #msg_content = msg_content.encode('utf-8')
                
                # 稍后解析出邮件:
                msg = Parser().parsestr(msg_content)
                #第一层编码错误
                text = Che.get_mail_text(msg)
                #print(text)
                #print(CheckCustomerConfirmExcelEmail().get_base_sw_name_and_config_name(text))
                #print(CheckCustomerConfirmExcelEmail().get_oder_num_and_board(text))


                """
                resp, lines, octets = server.retr(index)
                # lines存储了邮件的原始文本的每一行
                # 邮件的原始文本:
                msg_content = b'\r\n'.join(lines).decode('utf-8','ignore')
                print("msg_content",msg_content)
                msg = Parser().parsestr(msg_content)
                print("msg",msg)
                
                Che.get_att(msg)
                print("Che.get_att(msg)",Che.get_att(msg))
                """
                print("\n\n-------------------------------一封邮件处理完毕-------------------------------\n\n")
                subject_str = Che.decode_str(msg.get('Subject',''))
                print("subject_str",subject_str)
                if Che.subject_title in str(subject_str):
                    print("all_version",Che.get_base_sw_name_and_config_name(text))
                    tmp_basesw_version = Che.get_base_sw_name(text)#get_base_sw_name_and_config_name(text)[0]
                    if "MT9602" in str(subject_str):
                        basesw_version = Che.get_vidaa_version(str(tmp_basesw_version))   
                    else:
                        basesw_version = Che.get_just_sw_number(str(tmp_basesw_version))                    
                    print("basesw_version",basesw_version)
                    tmp_configsw_version = Che.get_config_sw_name(text)
                    configsw_version = Che.get_just_sw_number(str(tmp_configsw_version))
                    print("configsw_version",configsw_version)
                    order_info = Che.get_oder_num(text)
                    print("order_info",order_info)
                    fangan_info = Che.get_board_type(text)
                    print("fangan_info",fangan_info)
                    sum_confirm_email_once_time = sum_confirm_email_once_time + 1
                    confirm_line = confirm_line + 1
                    print("confirm_line",confirm_line)
                    sheet.cell(row=confirm_line, column=2).value = str(order_info)
                    sheet.cell(row=confirm_line, column=4).value = str(basesw_version)
                    sheet.cell(row=confirm_line, column=3).value = fangan_info
                    sheet.cell(row=confirm_line, column=5).value = configsw_version
                    #sheet.append(row=confirm_line)
                    #fr_confirm.save(Che.save_confirm_sw_file+'客户回复确认表格.xlsx')
                    print("\n\n-------------------------------获取邮件主题处理完毕-------------------------------\n\n")
                    Workbook.save(execl_path)
            confirm_line = 0
            sum_confirm_email_once_time = 0    
            Workbook.save(execl_path)
            #关闭连接
            server.quit()
    else:
        #记录文件不存在
        #第一次运行,创建并写入文件
        f = open(Che.counttxtpath, 'w',encoding='utf-8',errors='ignore')
        f.write(str(new_email_num))
        f.close()
