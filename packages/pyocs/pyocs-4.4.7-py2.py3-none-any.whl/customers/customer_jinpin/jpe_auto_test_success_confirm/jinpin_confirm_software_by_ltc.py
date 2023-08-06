import os
import logging
import time
import sys
print(sys.path)
sys.path.append(".")
import ocs_sendemail
import My_pyocs_fun
from openpyxl import load_workbook
from pyocs import pyocs_software
from pyocs.pyocs_demand import PyocsDemand
# 创建对象
import requests
import json
import ltc_interface
LTC = ltc_interface.LtcInterface()




class Logger(object):
    def __init__(self, filename='confirm_byocs.log', stream=sys.stdout):
        self.terminal = stream
        self.log = open(filename, 'a')

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        pass

sys.stdout = Logger(stream=sys.stdout)
sys.stdout = Logger(stream=sys.stderr)

sys_time=time.strftime('time:%Y-%m-%d-%H-%M-%S', time.localtime(time.time()))
print(sys_time)



# 打开Excel文件读取数据
dirPath = os.getcwd()
execl_path=dirPath+'/customers/customer_jinpin/jpe_auto_test_success_confirm/客户回复确认表格.xlsx'             
Workbook = load_workbook(execl_path)
# 获取行列等单元值
sheet = Workbook.get_sheet_by_name('test')
max_rows_new = sheet.max_row

row_index =0
while row_index <= max_rows_new:
    #我们自己的只有三列的表格
    #order_info 摘要信息既评审单号
    row_index = row_index +1
    order_info = sheet.cell(row=row_index, column=2)
    sw_info = sheet.cell(row=row_index, column=4)
    whole_sw_name = sheet.cell(row=row_index, column=5)
    fangan_info = sheet.cell(row=row_index, column=3)
    #记录方案信息和软件不为空的订单
    if None != order_info.value and str(sw_info.value).rstrip(" ") and '9602' not in fangan_info.value:
        customer_order_info = order_info.value
        print("customer_order_info",customer_order_info)
        sw_confirm_info = str(sw_info.value)
        print("sw_confirm_info",sw_confirm_info)

        find_order_payload = {
            "searchParams": [{
                "join": "和",
                "maxValue": "",
                "minValue": "",
                "operator": "包含",
                "conditionValue": [
                    customer_order_info
                ],
                "condition": "客户批号",
                "leftParenthesis": "0",
                "rightParenthesis": "0",
                "leftSquareBracket": "0",
                "rightSquareBracket": "0"
            },{
                "join": "和",
                "maxValue": "",
                "minValue": "",
                "operator": "在",
                "conditionValue": [
                    'TV研发组织'
                ],
                "condition": "申请研发组织",
                "leftParenthesis": "0",
                "rightParenthesis": "0",
                "leftSquareBracket": "0",
                "rightSquareBracket": "0"
            },{
                "join": "和",
                "maxValue": "",
                "minValue": "",
                "operator": "不包含",
                "conditionValue": [
                    '意向需求'
                ],
                "condition": "分类",
                "leftParenthesis": "0",
                "rightParenthesis": "0",
                "leftSquareBracket": "0",
                "rightSquareBracket": "0"
            }]
        }



        #1、获取待确认软件的订单
        #print("find_order_payload",find_order_payload)
        order_list = list()
        find_order_info_response = LTC.ltc_search_order_by_advance(find_order_payload)
        print("find_order_info_response.status_code",find_order_info_response.status_code)
        if find_order_info_response.status_code == 200:
            order_list = find_order_info_response.json().get("data")
        
        if order_list != None:
            for i in range(len(order_list)):
                ##order_list[i]= 'ST20801716'
                sw_response = LTC.ltc_get_order_software(order_list[i])
                print(sw_response.json())
                if sw_response.status_code != 200 or sw_response.json().get("data") == None:
                    break
                #订单里软件为空。
                sw_list = sw_response.json().get("data")
                already_confirmed = False
                if sw_list != None:
                    print(sw_response.json().get("data"))
                    for j in range(len(sw_list)):
                        
                        if sw_list[j].get("confirmed") == '邮件已确认' or sw_list[j].get("confirmed") == '无需确认':
                            already_confirmed = True
                            print("已确认",order_list[i])
                            break
                    if already_confirmed != True:
                        for j in range(len(sw_list)):
                            #if sw_list[j].get("number") == find_fw_id:#比较搜索到的固件订单上的是否匹配，不匹配的话就引用。
                            comfirm_response = LTC.ltc_set_confirm_by_email(order_list[i],sw_list[j].get("number"),'N',cus_confirm_info=order_list[i]+' ' +customer_order_info,sw_confirm_info=sw_confirm_info)
                            if comfirm_response.status_code == 200 and comfirm_response.json().get("message") == 'success':
                                    print("确认成功",order_list[i])
                                    break
                        
##after confirm delete data
#del Workbook['ted']
Workbook.remove(sheet)
Workbook.create_sheet('test')
Workbook.save(execl_path)

#response = LTC.ltc_set_confirm_by_email('ST20794039','FW221310977461','N',cus_confirm_info='ST20794039',sw_confirm_info='20221011_200612')

#print(response.json())
