import os
import logging
import time
import sys
print(sys.path)
sys.path.append(".")
import ocs_sendemail
import My_pyocs_fun
from openpyxl import load_workbook
from pyocs import pyocs_software
from pyocs.pyocs_demand import PyocsDemand
# 创建对象
#confirm_task = pyocs_software.PyocsSoftware()
confirm_task = My_pyocs_fun.child_pyocs()
send_email = ocs_sendemail.my_osc_sendemail()


# 打开Excel文件读取数据
dirPath = os.getcwd()
execl_path=dirPath+'\\customers\\customer_jinpin\\test_success_confirm\\客户回复确认表格.xlsx'             
Workbook = load_workbook(execl_path)
# 获取行列等单元值
sheet = Workbook.active
max_rows_new = sheet.max_row

to_engyneer = ['linxiangna@cvte.com','chenchaoxiong@cvte.com','liyangqiu@cvte.com','hejian@cvte.com']
#to_engyneer = ['linxiangna@cvte.com']

# main
row_index = 2
ocs_number = 000000
soft_version = str()
order_info = str()
fangan_info = str()
order_soft_success_list = list()
order_soft_fail_list = list()
chenchaoxiong_unpass_order = list()
linxiangna_unpass_order = list()
chengfan_unpass_order = list()
zhanghonghai_unpass_order = list()

chenchaoxiong_upload_ocs_sucess_order = list()
linxiangna_upload_ocs_sucess_order = list()
chengfan_upload_ocs_sucess_order = list()
zhanghonghai_upload_ocs_sucess_order = list()
upload_soft_to_ocs_success_list = list()
text_Enter = '\n'
text_stop_used = "软件被停用"

ocs_number = confirm_task.get_ocs_number_from_sw("20201006_095026", "01-202008-629")
print("ocs_number:",ocs_number)
