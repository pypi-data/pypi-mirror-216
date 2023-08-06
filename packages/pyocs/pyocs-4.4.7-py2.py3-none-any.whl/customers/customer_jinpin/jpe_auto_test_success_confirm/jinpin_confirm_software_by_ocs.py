import os
import logging
import time
import sys
print(sys.path)
sys.path.append(".")
import ocs_sendemail
import My_pyocs_fun
from openpyxl import load_workbook
from pyocs import pyocs_software
from pyocs.pyocs_demand import PyocsDemand
# 创建对象
#confirm_task = pyocs_software.PyocsSoftware()
confirm_task = My_pyocs_fun.child_pyocs()
send_email = ocs_sendemail.my_osc_sendemail()


# 打开Excel文件读取数据
dirPath = os.getcwd()
execl_path=dirPath+'/customers/customer_jinpin/jpe_auto_test_success_confirm/客户回复确认表格.xlsx'             
Workbook = load_workbook(execl_path)
# 获取行列等单元值
sheet = Workbook.active
max_rows_new = sheet.max_row

to_engyneer = ['linxiangna@cvte.com','chenchaoxiong@cvte.com','leimingsheng@cvte.com']
#to_engyneer = ['linxiangna@cvte.com']

# main
row_index = 1
ocs_number = 000000
soft_version = str()
order_info = str()
fangan_info = str()
order_soft_success_list = list()
order_soft_fail_list = list()
tmp_order_soft_fail_list = list()
chenchaoxiong_unpass_order = list()
linxiangna_unpass_order = list()
chengfan_unpass_order = list()
chenjiayi8469_unpass_order = list()

chenchaoxiong_upload_ocs_sucess_order = list()
linxiangna_upload_ocs_sucess_order = list()
chengfan_upload_ocs_sucess_order = list()
chenjiayi8469_upload_ocs_sucess_order = list()
upload_soft_to_ocs_success_list = list()
text_Enter = '\n'
text_stop_used = "软件被停用"



class Logger(object):
    def __init__(self, filename='confirm_byocs.log', stream=sys.stdout):
        self.terminal = stream
        self.log = open(filename, 'a')

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        pass

sys.stdout = Logger(stream=sys.stdout)
sys.stdout = Logger(stream=sys.stderr)

sys_time=time.strftime('time:%Y-%m-%d-%H-%M-%S', time.localtime(time.time()))
print(sys_time)



while row_index <= max_rows_new:
    #我们自己的只有三列的表格
    #order_info 摘要信息既评审单号
    #soft_version = sheet.cell(row=row_index, column=3)
    #fangan_info = sheet.cell(row=row_index, column=2)
    #DQA回复的表格
    order_info = sheet.cell(row=row_index, column=2)
    soft_version = sheet.cell(row=row_index, column=4)
    fangan_info = sheet.cell(row=row_index, column=3)
    #记录方案信息和软件不为空的订单
    if None != order_info.value and str(soft_version.value).rstrip(" "):
        tmp_success_list = [order_info.value, soft_version.value,fangan_info.value]
        order_soft_success_list.append(tmp_success_list)
    else:#找出评审单号为空的确认不通过的订单
         #评审单单号为空就啥都确认不了
        if(None != fangan_info.value):
            tmp_fail_list = [order_info.value, soft_version.value, fangan_info.value]
            #print("评审单号为空：",row_index)
            order_soft_fail_list.append(tmp_fail_list)

    row_index += 1
print(max_rows_new)
print(order_soft_success_list)
print(order_soft_fail_list)


##get_enable_software_list这个接口获取的信息太不对了
#先关联一遍
for key in order_soft_success_list:
    #意向订单，虚拟订单相同评审单号，相同软件时间版本会影响确认.现在很多都是直接在意向单上确认了，所以大货的没确认
    #key[0]来搜索的是订单和虚拟订单
    #只跳过了关联意向单的环节，确认环节，相同评审单的情况还保留，因为不好把握那个才是真正需要测的软件
    #这样的话，有可能原来订单有软件也关联了。应该是先排除一波这个条件

    # 根据sw_name和order_info，确认软件
    #能否改为根据未确认的订单列表来确认
    print("key",key)
    ocs_numlist=confirm_task.get_ocs_number_from_customer_order('金品',str(key[0]))
    print('ocs_numlist',ocs_numlist)
    #for循环执行了两次，所以才会上传两次软件
    #for dst_ocs_num in ocs_numlist:
    for i in range(0,len(ocs_numlist[1])) :
        is_success =0 
        dst_ocs_num=str(ocs_numlist[1][i])
        #dst_ocs_num = dst_ocs_num[2:-2]
        dst_ocs_status=str(ocs_numlist[0][i])
        #dst_ocs_status = dst_ocs_status[2:-2]
        print('dst_ocs_status',dst_ocs_status)
        print('dst_ocs_num',dst_ocs_num)
        #ocs_num_array=confirm_task.get_all_ocs_number_from_sw(str(key[1]))
        #print('ocs_num_array',ocs_num_array)
        #if dst_ocs_num not in ocs_num_array :
        #嘉艺的多是建虚拟单，没必要再关联软件，直接在虚拟单贴确认
        #if '706' in str(key[2]):
            #continue
        if dst_ocs_status != '已完成' and dst_ocs_status != '取消任务' and dst_ocs_status != '软件测试不通过':
            #判断ocs里有对应的软件就不引用了
            softlist = confirm_task.get_enable_software_list(str(dst_ocs_num))
            ##为啥里面没有实际的对应的软件也能获取到非空值，之前有的软件已经被删除了
            if(softlist!=None):
                print("softlist",softlist[0].name)
            #为啥会重复关联
            if  (softlist == None) or (str(key[1]) not in str(softlist[0].name)):
                print("do reuse job")
                is_success = confirm_task.reuse_old_sw_from_src_to_dst_by_fw_id(str(key[1]),str(dst_ocs_num), os.getcwd())

#避免延时不够，还没传完
time.sleep(180)

#先确认，确认失败，关联，再确认
#先关联一遍
for key in order_soft_success_list:
    #意向订单，虚拟订单相同评审单号，相同软件时间版本会影响确认.现在很多都是直接在意向单上确认了，所以大货的没确认
    #key[0]来搜索的是订单和虚拟订单
    #只跳过了关联意向单的环节，确认环节，相同评审单的情况还保留，因为不好把握那个才是真正需要测的软件
    #这样的话，有可能原来订单有软件也关联了。应该是先排除一波这个条件

    # 根据sw_name和order_info，确认软件
    #能否改为根据未确认的订单列表来确认
    #虚拟单怎么都没有了，新建的没有金品标识了
    ocs_numlist=confirm_task.get_ocs_number_from_customer_order('金品',str(key[0]))

    print('ocs_numlist',ocs_numlist)
    #for循环执行了两次，所以才会上传两次软件
    #for dst_ocs_num in ocs_numlist:
    for i in range(0,len(ocs_numlist[1])) :
        is_success =0 
        dst_ocs_num=str(ocs_numlist[1][i])
        #dst_ocs_num = dst_ocs_num[2:-2]
        dst_ocs_status=str(ocs_numlist[0][i])
        #dst_ocs_status = dst_ocs_status[2:-2]
        print('dst_ocs_status',dst_ocs_status)
        print('dst_ocs_num',dst_ocs_num)
        if dst_ocs_status != '已完成' and dst_ocs_status != '取消任务' and dst_ocs_status != '软件测试不通过':
            #判断ocs里有对应的软件就不引用了
            ##这里为啥获取的又是空的，再运行一次就又有了
            softlist = confirm_task.get_enable_software_list(str(dst_ocs_num))
            print("confirm softlist",softlist)
            #if(softlist!=None):
                #print("confirm softlist swname",softlist[0].name)
            #if  (softlist != None) and (str(key[1]) in str(softlist[0].name)):
            #print("confirm softlist swname",softlist[0].name)
            ret = confirm_task.set_enable_software_confirm_for_ocs(str(dst_ocs_num),str(key[1]),4)
            if int(ret) == -3:
                #好多--3的返回值，多是目标订单没有对应的软件。直接重新确认一遍。
                print("confirm again")
                ret = confirm_task.set_enable_software_confirm_for_ocs(str(dst_ocs_num),str(key[1]),4)
            print("ret",ret)

