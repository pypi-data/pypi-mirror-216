import os
import logging
import time
import sys
print(sys.path)
sys.path.append(".")
import ocs_sendemail
import My_pyocs_fun
from openpyxl import load_workbook
from pyocs import pyocs_software
from pyocs.pyocs_demand import PyocsDemand
# 创建对象
#confirm_task = pyocs_software.PyocsSoftware()
confirm_task = My_pyocs_fun.child_pyocs()
send_email = ocs_sendemail.my_osc_sendemail()


# 打开Excel文件读取数据
dirPath = os.getcwd()
execl_path=dirPath+'/customers/customer_jinpin/test_success_confirm/客户回复确认表格.xlsx'             
Workbook = load_workbook(execl_path)
# 获取行列等单元值
sheet = Workbook.active
max_rows_new = sheet.max_row

#to_engyneer = ['linxiangna@cvte.com','chenchaoxiong@cvte.com','liyangqiu@cvte.com','chenjiayi8469@cvte.com']
to_engyneer = ['linxiangna@cvte.com']

# main
row_index = 2
ocs_number = 000000
soft_version = str()
order_info = str()
fangan_info = str()
order_soft_success_list = list()
order_soft_fail_list = list()
tmp_order_soft_fail_list = list()
chenchaoxiong_unpass_order = list()
linxiangna_unpass_order = list()
chengfan_unpass_order = list()
zhanghonghai_unpass_order = list()

chenchaoxiong_upload_ocs_sucess_order = list()
linxiangna_upload_ocs_sucess_order = list()
chengfan_upload_ocs_sucess_order = list()
zhanghonghai_upload_ocs_sucess_order = list()
upload_soft_to_ocs_success_list = list()
text_Enter = '\n'
text_stop_used = "软件被停用"



class Logger(object):
    def __init__(self, filename='confirm.log', stream=sys.stdout):
        self.terminal = stream
        self.log = open(filename, 'a')

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        pass

sys.stdout = Logger(stream=sys.stdout)
sys.stdout = Logger(stream=sys.stderr)

sys_time=time.strftime('time:%Y-%m-%d-%H-%M-%S', time.localtime(time.time()))
print(sys_time)



while row_index <= max_rows_new:
    #我们自己的只有三列的表格
    #order_info 摘要信息既评审单号
    #soft_version = sheet.cell(row=row_index, column=3)
    #fangan_info = sheet.cell(row=row_index, column=2)
    #DQA回复的表格
    order_info = sheet.cell(row=row_index, column=2)
    soft_version = sheet.cell(row=row_index, column=4)
    fangan_info = sheet.cell(row=row_index, column=3)
    #记录方案信息和软件不为空的订单
    if None != order_info.value and str(soft_version.value).rstrip(" "):
        tmp_success_list = [order_info.value, soft_version.value,fangan_info.value]
        order_soft_success_list.append(tmp_success_list)
    else:#找出评审单号为空的确认不通过的订单
         #评审单单号为空就啥都确认不了
        if(None != fangan_info.value):
            tmp_fail_list = [order_info.value, soft_version.value, fangan_info.value]
            #print("评审单号为空：",row_index)
            order_soft_fail_list.append(tmp_fail_list)

    row_index += 1
print(max_rows_new)
print(order_soft_success_list)
print("failed")
print(order_soft_fail_list)

"""
for key in order_soft_success_list:
    #意向订单，虚拟订单相同评审单号，相同软件时间版本会影响确认
    #key[0]来搜索的是订单和虚拟订单
    #只跳过了关联意向单的环节，确认环节，相同评审单的情况还保留，因为不好把握那个才是真正需要测的软件
    #这样的话，有可能原来订单有软件也关联了。应该是先排除一波这个条件

    # 根据sw_name和order_info，确认软件

     
    if((str(key[0]) !=None) and  (str(key[1]) !=None)):  
        sw_name = str(key[1])
        order_info =  str(key[0])  
        ocs_number = self.get_ocs_number_from_sw(sw_name, order_info)
        print("can find  ocs_number",ocs_number)
        if ocs_number is None:
            pass
        else:
            sw_list = self.get_enable_software_list(ocs_number=ocs_number)
    if(False == confirm_task.write_order_have_deal(str(key[0]))):
        continue
    #订单上已有可以确认的软件则跳过，多个怎么知道那个是呢，还会分先后
    #先不关联，先确认，确认失败了再看这个订单还有没有其他软件版本，有的话先关联
    else if((sw_list !=None) and (order_info in sw_list)):
        continue
    else:
        ocs_numlist=confirm_task.get_ocs_number_from_customer_order('金品',str(key[0]))
        print('ocs_numlist',ocs_numlist)
        #for循环执行了两次，所以才会上传两次软件
        #for dst_ocs_num in ocs_numlist:
        for i in range(0,len(ocs_numlist[1])) :
            dst_ocs_num=str(ocs_numlist[1][i])
            #dst_ocs_num = dst_ocs_num[2:-2]
            dst_ocs_status=str(ocs_numlist[0][i])
            #dst_ocs_status = dst_ocs_status[2:-2]
            print('dst_ocs_status',dst_ocs_status)
            print('dst_ocs_num',dst_ocs_num)
            #ocs_num_array=confirm_task.get_all_ocs_number_from_sw(str(key[1]))
            #print('ocs_num_array',ocs_num_array)
            #if dst_ocs_num not in ocs_num_array :
            #嘉艺的多是建虚拟单，没必要再关联软件，直接在虚拟单贴确认
            if '706' in str(key[2]):
                continue
            if dst_ocs_status != '已完成' and dst_ocs_status != '取消任务' and dst_ocs_status != '软件测试不通过':
                    is_success = confirm_task.reuse_old_sw_from_src_to_dst_by_fw_id(str(key[1]),str(dst_ocs_num), os.getcwd())

time.sleep(60)
"""
for key in order_soft_success_list:
    #ret = confirm_task.set_software_confirm(key[1], key[0])
    ret = confirm_task.set_software_confirm(key[1], key[0])
    ocs_number = confirm_task.get_ocs_number_from_sw(key[1], key[0])
    key.append(ocs_number)
    if ret:
        print(key[0] + " -- " + key[1] + " confirm success")
        upload_soft_to_ocs_success_list.append(key)
    else:#找出评审单位不为空的确认不通过的订单
        tmp_fail_list = key
        tmp_order_soft_fail_list.append(tmp_fail_list)

for key in tmp_order_soft_fail_list:
    #这里获取的都是大货订单或者虚拟订单的ocsnum
    ocs_numlist=confirm_task.get_ocs_number_from_customer_order('金品',str(key[0]))
    print('ocs_numlist',ocs_numlist)
    #for循环执行了两次，所以才会上传两次软件
    #for dst_ocs_num in ocs_numlist:
    for i in range(0,len(ocs_numlist[1])) :
        dst_ocs_num=str(ocs_numlist[1][i])
        #dst_ocs_num = dst_ocs_num[2:-2]
        dst_ocs_status=str(ocs_numlist[0][i])
        #dst_ocs_status = dst_ocs_status[2:-2]
        print('dst_ocs_status',dst_ocs_status)
        print('dst_ocs_num',dst_ocs_num)
        #ocs_num_array=confirm_task.get_all_ocs_number_from_sw(str(key[1]))
        #print('ocs_num_array',ocs_num_array)
        #if dst_ocs_num not in ocs_num_array :
        #嘉艺的多是建虚拟单，没必要再关联软件，直接在虚拟单贴确认
        if '706' in str(key[2]):
            continue
        if dst_ocs_status != '已完成' and dst_ocs_status != '取消任务' and dst_ocs_status != '软件测试不通过':
            #摘要在确认通过的订单里存在也不做引用
            #if(str(key[0]) in upload_soft_to_ocs_success_list[0]):
                #key.append(str('该评审单已有确认'))
                #upload_soft_to_ocs_success_list.append(key)
            #else:
                #怎么都无法重复运行确认了
            if(False == confirm_task.write_order_have_deal(str(key[0]))):
                continue 
            else:   
                is_success = confirm_task.reuse_old_sw_from_src_to_dst_by_fw_id(str(key[1]),str(dst_ocs_num), os.getcwd())        
    #确认不成功的重新上传再确认一次，这次还不成功的才判为不成功
    ret = confirm_task.set_software_confirm(key[1], key[0])
    ocs_number = confirm_task.get_ocs_number_from_sw(key[1], key[0])
    key.append(ocs_number)
    if ocs_number != None:
        key.append(str('推测被禁用'))
    if ret:
        print(key[0] + " -- " + key[1] + " confirm success")
        upload_soft_to_ocs_success_list.append(key)
    else:
        #这其中又一波是自建虚拟单，和原来评审单哈不一样的
        tmp_fail_list = key 
        order_soft_fail_list.append(tmp_fail_list)
print("key in order_soft_success_list")
print(key)

"""
for key in order_soft_success_list:
    link=confirm_task.get_audit_failed_link(key[3], key[2])
    print(link)
    if link:
        key[3]=link
"""
"""
for key in order_soft_success_list:
    pd = PyocsDemand(ocs_number)
    html=pd.get_ocs_html()
    print(html)
    orderlist=ocs_communication.get_enable_software_list_from_html(html)
"""

"""
#查找
for key in order_soft_fail_list:
    #意向订单，虚拟订单相同评审单号，相同软件时间版本会影响确认
    #key[0]来搜索的是订单和虚拟订单
    ocs_numlist=confirm_task.get_ocs_number_from_customer_order('金品',str(key[0]))
    print('ocs_numlist',ocs_numlist)
    #for循环执行了两次，所以才会上传两次软件
    #for dst_ocs_num in ocs_numlist:
    for i in range(0,len(ocs_numlist[1])) :
        dst_ocs_num=str(ocs_numlist[1][i])
        #dst_ocs_num = dst_ocs_num[2:-2]
        dst_ocs_status=str(ocs_numlist[0][i])
        #dst_ocs_status = dst_ocs_status[2:-2]
        print('dst_ocs_status',dst_ocs_status)
        print('dst_ocs_num',dst_ocs_num)
        #ocs_num_array=confirm_task.get_all_ocs_number_from_sw(str(key[1]))
        #print('ocs_num_array',ocs_num_array)
        #if dst_ocs_num not in ocs_num_array :
        if dst_ocs_status != '已完成':
            is_success = confirm_task.reuse_old_sw_from_src_to_dst_by_fw_id(str(key[1]),str(dst_ocs_num), os.getcwd())

time.sleep(60)
"""


for key in upload_soft_to_ocs_success_list:
    #根据方案找upload_soft_to_ocs_success_list的对应工程师
    engineer_name=ocs_sendemail.my_osc_sendemail.get_engineer_name(key[2])
    tmp_fail_list=key
    if("chenchaoxiong@cvte.com"==engineer_name):
        chenchaoxiong_upload_ocs_sucess_order.append(tmp_fail_list)
    if ("linxiangna@cvte.com" == engineer_name):
        linxiangna_upload_ocs_sucess_order.append(tmp_fail_list)
    if ("hejian@cvte.com" == engineer_name):
        chengfan_upload_ocs_sucess_order.append(tmp_fail_list)
    if ("zhanghonghai@cvte.com" == engineer_name):
        zhanghonghai_upload_ocs_sucess_order.append(tmp_fail_list)

for key in order_soft_fail_list:
    #根据方案找order_soft_fail_list的对应工程师
    engineer_name=ocs_sendemail.my_osc_sendemail.get_engineer_name(key[2])
    tmp_fail_list=key
    print(engineer_name)
    if("chenchaoxiong@cvte.com"==engineer_name):
        chenchaoxiong_unpass_order.append(tmp_fail_list)
    if ("linxiangna@cvte.com" == engineer_name):
        linxiangna_unpass_order.append(tmp_fail_list)
    if ("hejian@cvte.com" == engineer_name):
        chengfan_unpass_order.append(tmp_fail_list)
    if ("zhanghonghai@cvte.com" == engineer_name):
        zhanghonghai_unpass_order.append(tmp_fail_list)

#if(0 != len(order_soft_fail_list)):
new_str_chenfang = send_email.failed_list_to_str(chengfan_unpass_order)
new_str_chenchaoxiong = send_email.failed_list_to_str(chenchaoxiong_unpass_order)
new_str_xiangna = send_email.failed_list_to_str(linxiangna_unpass_order)
new_str_honghai = send_email.failed_list_to_str(zhanghonghai_unpass_order)
upload_ocs_sucess_str_chenfang = send_email.failed_list_to_str(chengfan_upload_ocs_sucess_order)
upload_ocs_sucess_str_chenchaoxiong = send_email.failed_list_to_str(chenchaoxiong_upload_ocs_sucess_order)
upload_ocs_sucess_str_xiangna = send_email.failed_list_to_str(linxiangna_upload_ocs_sucess_order)
upload_ocs_sucess_str_honghai = send_email.failed_list_to_str(zhanghonghai_upload_ocs_sucess_order)
print(new_str_chenchaoxiong)
print(upload_ocs_sucess_str_chenchaoxiong)
confirm_success_text = "亲爱的工程师，您好，以下订单自动确认通过"
upload_ocs_sucess_text_chenfang = text_Enter + "嘉艺:" + text_Enter + upload_ocs_sucess_str_chenfang
upload_ocs_sucess_text_chenchaoxiong = text_Enter + "潮雄:" + text_Enter + upload_ocs_sucess_str_chenchaoxiong
upload_ocs_sucess_text_xiangna = text_Enter + "祥纳:" + text_Enter + upload_ocs_sucess_str_xiangna
upload_ocs_sucess_text_honghai = text_Enter + "宏海:" + text_Enter + upload_ocs_sucess_str_honghai
confirm_success_list = confirm_success_text+upload_ocs_sucess_text_chenfang + upload_ocs_sucess_text_chenchaoxiong + upload_ocs_sucess_text_xiangna + upload_ocs_sucess_text_honghai
#confirm_success_list = confirm_success_text+text_Enter+send_email.failed_list_to_str(upload_soft_to_ocs_success_list)+ text_Enter
text = "亲爱的工程师，您好，以下订单无法自动确认通过，请及时关注"
text_chenfang = text_Enter + "嘉艺:" + text_Enter + new_str_chenfang
text_chenchaoxiong = text_Enter + "潮雄:" + text_Enter + new_str_chenchaoxiong
text_xiangna = text_Enter + "祥纳:" + text_Enter + new_str_xiangna
text_honghai = text_Enter + "宏海:" + text_Enter + new_str_honghai

suggest_info_text =text_Enter +  "亲爱的各位工程师记得及时梳理确认后测试类型为不用测试的订单 http://ocs.gz.cvte.cn/tv/Tasks/index/SearchId:2950404/range:my"
suggest_to_test_text =text_Enter +  "以下订单需注意设置恰当测试类型并提醒补测 https://ocs.gz.cvte.cn/tv/Tasks/index/SearchId:2950510/range:my"
whole_text = confirm_success_list+text + text_chenfang + text_chenchaoxiong + text_xiangna  + suggest_info_text + suggest_to_test_text
print(whole_text)
send_email.send_email_to_engineer(to_engyneer,whole_text)
