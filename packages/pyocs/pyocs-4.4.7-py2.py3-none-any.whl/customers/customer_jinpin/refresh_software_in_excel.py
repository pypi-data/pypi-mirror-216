import xlrd
import json
import requests
from pyocs import pyocs_software
from openpyxl import load_workbook
import xlwt

class SoftwareDownloadlinkLtc:
    def __init__(self):
        self.name = ''  # 软件名
        self.download_link = ''  # 软件下载链接
        self.deadline = ''  # 链接截止日期
"""根据软件局部信息更新软件外网下载链接和链接更新日期
Args:
    firmware_name: 软件名查找
    firmware_version: 通过固件版本查找。
    firmware_version和firmware_name可以单独使用，也可以合在一起使用。单独使用范围更广。合在一起使用更精确。
    check_info 限制包名必须包含什么内容
    need_emmc_info 是否需要带上emmc的信息 ，这个有bug，暂时捞不出来
Returns:
    返回通过该信息拿到的  所有   软件刷新后的下载链接和下载链接有效期，对象列表
"""
def refreshsw_by_ltc(firmware_name='',firmware_version='',check_info = '' ,need_emmc_info = False):
    url = "https://faas.gz.cvte.cn/function/cplm-refresh-software-link"

    payload = {
    "firmware_id": "",
    "firmware_name": firmware_name,
    "firmware_version": firmware_version
    }
    headers = {
    'Content-Type': 'application/json',
    'Cookie': 'BIGipServerpool_yp_faas_80=1427180042.20480.0000'
    }

    response = requests.request("POST", url, headers=headers, json=payload)
    #print(response.json().get("data"))
    sdl_list = list()
    if response.status_code == 200 and response.json().get("data")!=None:
        print(response.json())
        sdl = SoftwareDownloadlinkLtc()

        if need_emmc_info == False:
            for i in  range(len(response.json().get("data"))):

                if 'EMMCBIN_' not in response.json().get("data")[i].get("software_name") and check_info in response.json().get("data")[i].get("software_name"):
                    sdl.name = response.json().get("data")[i].get("software_name")  # 软件名
                    sdl.download_link = response.json().get("data")[i].get("software_link")  # 软件下载链接
                    sdl.deadline = response.json().get("data")[i].get("expired_time")  # 链接截止日期
                sdl_list.append(sdl)
            return sdl_list
        else:
            for i in  range(len(response.json().get("data"))):
                if  check_info in response.json().get("data")[i].get("software_name"):
                    sdl.name = response.json().get("data")[i].get("software_name")  # 软件名
                    sdl.download_link = response.json().get("data")[i].get("software_link")  # 软件下载链接
                    sdl.deadline = response.json().get("data")[i].get("expired_time")  # 链接截止日期
                sdl_list.append(sdl)
            return sdl_list
        
    else:
        return sdl_list

path = 'SW_REFRESH.xlsx'#代码运行目录。自己可以改路径。
wb = load_workbook(path)    #cplm 上导出 按钮，的文件，放在代码路径下

if __name__ == '__main__':

    sheet = wb.active


    refresh_task = pyocs_software.PyocsSoftware()
    
    for row_index in range(2,sheet.max_row):
        if sheet.cell(row=row_index, column=1).value == None:
            continue
        if len(sheet.cell(row=row_index, column=1).value)<5:
            continue
        # thistwo case only for jpe    
        if ')' in sheet.cell(row=row_index, column=1).value.strip():
            sw_info = sheet.cell(row=row_index, column=1).value.strip().split(')')[1]
        elif '）' in sheet.cell(row=row_index, column=1).value.strip():
            sw_info = sheet.cell(row=row_index, column=1).value.strip().split('）')[1]
        else:
            sw_info = sheet.cell(row=row_index, column=1).value.strip()
        if '.zip' not in sw_info:
            sw_info = sw_info + '.zip'
        print("sw_info",sw_info)
        #result = refresh_task.get_refresh_software_download_link_by_sw_info(sw_info)
        #result = refreshsw_by_ltc(sw_info)
        #这样都还找不到，有可能是包名不对了，改为只看版本来找
        sw_version = sw_info.rsplit('_',1)[1].split('.zip')[0].strip()
        check_info = "JPE"
        
        result = refreshsw_by_ltc(firmware_name='',firmware_version=sw_version,check_info = check_info ,need_emmc_info = True)
        print('result ', result)
        if not result:

            print(sw_info + "软件未找到，请确认!")
        else:
            for sw_download in result:
                print("刷新软件：" + sw_info)
                print("软件包名：" + sw_download.name)
                print("下载链接：" + sw_download.download_link + " 有效截止时间：" + sw_download.deadline + "\n")
                new_link = sw_download.name+ "\n"+ sw_download.download_link + "\n"+ " 有效截止时间：" + sw_download.deadline 
                sheet.cell(row=row_index, column=2).value = new_link
        wb.save('SW_REFRESH.xlsx')
    wb.save('SW_REFRESH.xlsx')   

#refreshsw_by_ltc('CP702038_JPE_VT0165UB1_0102_MT9602_PB801_KENYA_V650DJB_Q01_C1_IR_R45MMI804_Logovnp_PID22_7KEY_r110023_C_4K_PA_V0102.01.00L.L1228.zip')