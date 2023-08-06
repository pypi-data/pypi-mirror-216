import openpyxl
from NutDriver import download_file_form_drive

FAE_COLUMN = 'B'
FAE_ACCOUNT_COLUMN = 'C'
FAE_DEPARTMENT_COLUMN = 'A'

download_file_form_drive('FaeInfo.xlsx')
workbook = openpyxl.load_workbook("FaeInfo.xlsx")
name_list = workbook.sheetnames
sheet = workbook[name_list[0]]

def get_fae_list():
    ret = list()
    for i in range(2,sheet.max_row+1):
        ret.append(sheet[FAE_COLUMN+str(i)].value)
    return ret

def get_fae_account_list():
    ret = list()
    for i in range(2,sheet.max_row+1):
        ret.append(sheet[FAE_ACCOUNT_COLUMN+str(i)].value)
    return ret

def get_fae_department(fae_name):
    if fae_name not in get_fae_list():
        return ''
    department_list = get_department_list()
    for department in department_list:
        fae_list = get_department_fae_list(department)
        if fae_name in fae_list:
            return department
    return ''

def get_fae_manager(fae_name):
    fae_department = get_fae_department(fae_name)
    return get_department_manager(fae_department)

def get_department_list():
    ret = list()
    for i in list(sheet.merged_cells):
        ret.append(sheet[i.start_cell.coordinate].value)
    return ret

def get_department_fae_list(department):
    ret = list()
    department_list = get_department_list()
    for i, val in enumerate(department_list):
        if department != val:
            continue
        min_row = list(sheet.merged_cells)[i].min_row
        max_row = list(sheet.merged_cells)[i].max_row
        for i in range(min_row,max_row + 1):
            ret.append(sheet[FAE_COLUMN + str(i)].value)
    return ret

def get_department_manager(department):
    department_list = get_department_list()
    if department not in department_list:
        return ''
    for i, val in enumerate(department_list):
        if department != val:
            continue
        row = list(sheet.merged_cells)[i].start_cell.row
        return sheet[FAE_COLUMN + str(row)].value

def get_fae_account(fae_name):
    for i in range(2,sheet.max_row+1):
        if fae_name != sheet['B'+str(i)].value:
            continue
        return sheet['C'+str(i)].value

