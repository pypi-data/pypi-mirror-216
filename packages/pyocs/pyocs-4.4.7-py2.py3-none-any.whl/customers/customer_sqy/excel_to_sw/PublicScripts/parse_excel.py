import openpyxl
import re

#--------------------------------------------------------
#  EXCEL 解析
#--------------------------------------------------------
workbook=openpyxl.load_workbook("/ssd2/zhanghonghai/codes/pyocs/pyocs/customers/customer_sqy/excel_to_sw/111.xlsx")
sheets=workbook.get_sheet_names()
# print(sheets)
request_sheet=workbook.get_sheet_by_name(sheets[0])

request_rows=request_sheet.max_row
request_columns=request_sheet.max_column

ReuestInfo_key={  "order_id"            :"D2",    #启悦订单号
                  "def_country"         :"D3",    #国家
                  "board_type"          :"B4",    #板卡型号
                  "bom_description"     :"D4",    #物料描述
                  "bom_id"              :"B5",    #客料号
                  "panel_type"          :"B6",    #屏型号
                  "ir_type"             :"B7",    #遥控器
                  "sound_type"          :"B8",    #喇叭功率
                  "language_list"       :"B10",   #OSD语言
                  "def_language"        :"D10",   #默认语言
                  "cec_arc_text"        :"D11",   #CEC,ARC
                  "pvr_text"            :"B14",   #PVR,时移
                  "hotel_mode"          :"B15",   #酒店模式
                  "ci_type"             :"D14",   #CI
                  "keypad_type"         :"D15",   #按键板
                  "app_list"            :"B17",   #APK
                  "other_request"       :"B18",   #其他需求
                  }

def get_request_value_by_key(key):
    valuestr = request_sheet[ReuestInfo_key[key]].value
    valuestr = str(valuestr)
    if valuestr != '':
        return valuestr
    else:
        return "ERROR"


class SoftwareRequest(object):
    """customer software request"""
    def __init__(self):
        super(SoftwareRequest, self).__init__()

    #====================================
    # 获取客户订单号
    #====================================
    def get_order_num(self):
        raw_str = get_request_value_by_key('order_id')
        pattern="[\u4e00-\u9fa5]+"#匹配中文
        regex = re.compile(pattern)
        raw_str = re.sub(regex,'',raw_str)#去掉中文

        pattern_2="[\uFF00-\uFFEF]+"#匹配半角及全角字符
        regex_2 = re.compile(pattern_2)
        raw_str = re.sub(regex_2,'',raw_str)#去掉半角及全角字符

        if '/' in raw_str:
            order_num_major,order_num_minor = raw_str.split("/")
            if order_num_major != '' and order_num_minor != '':
                return (order_num_major + "_" + order_num_minor)
            else:
                return 'SAMPLE'
        else:
                return 'SAMPLE'

    #====================================
    # 获取客户订单客料号
    #====================================
    def get_bom_num(self):
        return (get_request_value_by_key('bom_id'))

    def get_board_type(self):
        raw_str = get_request_value_by_key("board_type")
        pattern="[\u4e00-\u9fa5]+"#匹配中文
        regex = re.compile(pattern)
        raw_str = re.sub(regex,'',raw_str)#去掉中文
        print(raw_str)

        text1,text2,text3 = raw_str.split('.')
        return (text2 + '_' + text3)

    def get_def_country(self):
        return get_request_value_by_key("def_country")

    def get_panel_type(self):
        raw_str = get_request_value_by_key("panel_type")
        pattern="[\u4e00-\u9fa5]+"#匹配中文
        regex = re.compile(pattern)
        raw_str = re.sub(regex,'',raw_str)#去掉中文

        pattern_2="[\uFF00-\uFFEF]+"#匹配半角及全角字符
        regex_2 = re.compile(pattern_2)
        raw_str = re.sub(regex_2,'',raw_str)#去掉半角及全角字符
        print("get_panel_type = " + raw_str + "\n")

        return raw_str

    def get_def_language(self):
        return (get_request_value_by_key("def_language"))

    def get_language_list(self):
        return (get_request_value_by_key("language_list"))

    def get_ir_type(self):
        return (get_request_value_by_key("ir_type"))

    def get_keypad_type(self):
        return (get_request_value_by_key("keypad_type"))

    def get_app_list(self):
        return (get_request_value_by_key("app_list"))

    def get_hotel_mode(self):
        return (get_request_value_by_key("hotel_mode"))

    def get_cec_text(self):
        return (get_request_value_by_key("cec_arc_text"))

    def get_arc_text(self):
        return (self.software_request[ReuestInfo_key['arc_text']])

    def get_ci_type(self):
        return (get_request_value_by_key("ci_type"))

    def get_special_needs(self):
        return (get_request_value_by_key("other_request"))

    


