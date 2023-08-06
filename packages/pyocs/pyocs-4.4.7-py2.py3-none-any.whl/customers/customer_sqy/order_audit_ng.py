import os
import sys
PROJ_HOME_PATH = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
FILE_PATH = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(1,PROJ_HOME_PATH)

import re
import json
import openpyxl

from FaeInfo import *

from w3lib import html
import requests

from pyocs.pyocs_list import PyocsList
from pyocs.pyocs_software import PyocsSoftware
from pyocs.pyocs_demand import PyocsDemand

def get_searchid():
    fae_list = get_fae_list()
    advanced_search = {
    "0":{"search_field_name":"Task.create_time","search_field_id":"585","search_field_type":"10",
        "search_field_rel_obj":"null","search_opr":"TDD_OPER_RECENT",
        "input1":"14","input2":'null',"offset":"TDD_OFFSET_DAY"},
    "1":{"search_field_name":"Task.status","search_field_id":"584","search_field_type":"19",
        "search_field_rel_obj":"Enums","search_opr":"TDD_OPER_NOT_INC",
        "input1":"待录入需求","input2":'null',"offset":'null'},
    "2":{"search_field_name":"Task.status","search_field_id":"584","search_field_type":"19",
        "search_field_rel_obj":"Enums","search_opr":"TDD_OPER_NOT_INC",
        "input1":"暂停任务","input2":'null',"offset":'null'},
    "3":{"search_field_name":"Task.status","search_field_id":"584","search_field_type":"19",
        "search_field_rel_obj":"Enums","search_opr":"TDD_OPER_NOT_INC",
        "input1":"取消任务","input2":'null',"offset":'null'},
    "4":{"search_field_name":"Task.subject","search_field_id":"554","search_field_type":"5",
        "search_field_rel_obj":"null","search_opr":"TDD_OPER_NOT_INC",
        "input1":"视睿","input2":'null',"offset":'null'},
    "5":{"search_field_name":"Task.sw_user_id","search_field_id":"555","search_field_type":"19",
        "search_field_rel_obj":"Users","search_opr":"TDD_OPER_INC",
        "input1":' '.join(fae_list),"input2":'null',"offset":'null'}
    }
    condition = "1 and 2 and 3 and 4 and 5 and 6"
    searchid = PyocsSoftware().get_searchid_from_advanced_search(advanced_search, condition)
    return searchid

def get_ocs_list(searchid):
    page = 1
    ocs_list = []
    while True:
        (count_tmp,list_tmp) = PyocsList().get_ocs_id_list(str(searchid) + "/page:" + str(page))
        if count_tmp == 0:
            break
        ocs_list = ocs_list + list_tmp
        page = page + 1
    return ocs_list

def write_to_xlsx_file(xls_file_path,content_dict_list):
    wb=openpyxl.load_workbook(xls_file_path)
    shts=wb.get_sheet_names()
    sht=wb.get_sheet_by_name(shts[0])
    rows=sht.max_row
    columns=sht.max_column
    for line in content_dict_list:
        sht.append(line)
    #设置单元格自动换行
    for i in range(2,sht.max_row+1):
        sht['E'+str(i)].alignment = openpyxl.styles.Alignment(wrapText=True)
    wb.save(xls_file_path)

def upload_to_nutstore(file_path: str):
    upload_nutstore_url = "https://drive.cvte.com/d/ajax/pubops/uploadXHR"

    form_data = dict()
    headers = dict()
    headers.update({
        'Origin': 'https://drive.cvte.com',
        'Referer': 'https://drive.cvte.com/p/Dby6kwUQ2eACGMiDDA',
        'Content-Type': 'application/octet-stream'
    })
    form_data.update({
        'key': 'Dby6kwUQ2eACGMiDDA',
        'dirName': '/',
        'path': '/' + file_path.split('/')[-1]
    })
    payload =  open(file_path, 'rb')
    r = requests.post(upload_nutstore_url, params=form_data,data=payload, headers=headers)
    return r.text

if __name__ == '__main__':
    searchid = get_searchid()
    print(searchid)
    ocs_list = get_ocs_list(searchid)
    # ocs_list = ['649308','649308']
    ret = list()
    for ocs_item in ocs_list:
        pd = PyocsDemand(ocs_item)
        order_audit = pd.get_software_audit()
        comment_str = html.remove_tags(pd.get_task_comment_area())
        order_engineer = pd.get_ocs_software_engineer()
        TaskCommentPattern = re.compile(r'"cmt_id.*?"type":"TaskComment"}', re.DOTALL)
        comment_list = TaskCommentPattern.findall(comment_str)
        # audit_ng_reason = list()
        index = 0
        for comment in comment_list:
            index = index + 1
            if '"user_name":"'+order_audit+'"' not in comment:
                continue
            if "审核不通过审核不通过原因" not in comment:
                continue
            ng_sw = re.findall('"comment":"(.*)审核不通过审核不通过原因',comment)[0]
            ng_time = re.findall('"update_time":(.*),"update_user_id"',comment)[0]
            ng_reason = re.findall('审核不通过审核不通过原因：(.*)","update_time"',comment)[0]
            # audit_ng_reason.append(str(index) + '、' + ng_reason + ';审核时间:' + ng_time)
            ret.append([ocs_item,order_engineer,order_audit,ng_sw,ng_reason + ';审核时间:' + ng_time])
    write_to_xlsx_file(FILE_PATH + '/审核NG订单统计.xlsx',ret)
    upload_to_nutstore(FILE_PATH + '/审核NG订单统计.xlsx')

