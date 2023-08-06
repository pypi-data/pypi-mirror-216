import datetime
import logging
import os,sys
from openpyxl import load_workbook
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # __file__获取执行文件相对路径，整行为取上一级的上一级目录
sys.path.append(BASE_DIR)
sys.path.append(".")
from pyocs.pyocs_request import PyocsRequest
from pyocs import pyocs_software
from lxml import etree
from pyocs.pyocs_demand import PyocsDemand
class Software:
    def __init__(self):
        self.name = ''  # 软件名
        self.attachment_id = ''  # 软件attachment id
        self.fw_id = ''  # 软件fw_id

class child_pyocs(pyocs_software.PyocsSoftware):
    dirPath = os.getcwd()+"/customers/customer_hanke/atuo_do_daice_sw/"
    _ocs_base_link = 'https://ocs.gz.cvte.cn'
    # _ocs_base_link = 'https://ocstest.gz.cvte.cn'
    _search_url = _ocs_base_link + '/tv/Attachments/search_firmwares'
    _ocs_link_Prefix = _ocs_base_link + '/tv/Tasks/view/'    
    print("dirPath",dirPath)
    def __init__(self):
        self._logger.setLevel(level=logging.WARNING)  # 控制打印级别

    def upload_old_sw_by_id(self, ocs_num, old_sw_id, burn_place_hold, burn_type, disable_origin_sw=True):
        """根据ocs_num 和old_sw_name， 上传库存软件
        """
        fields = {'data[fw_Attach][file][]': '(binary)',
                  'data[fw_Attach][old_fw_id][]': str(old_sw_id),
                  'data[fw_Attach][ab_sw_name][]': '',
                  'data[fw_Attach][ab_sw_id][]': '',
                  'data[fw_Attach][ab_xml_name][]': '',
                  'data[fw_Attach][ab_xml_id][]': '',
                  'data[fw_Attach][autobuild_bill_no][]': '',
                  'data[fw_Attach][fileXML][]': '(binary)',
                  'data[fw_Attach][test_type][]': "100",  # 5为E类测试,100为不用测试
                  'data[fw_Attach][fw_cfm_type]': '',
                  'data[fw_Attach][soft_remark]': '',
                  'data[fw_Attach][burn_place_hold][]': burn_place_hold,
                  'data[fw_Attach][burn_type]': burn_type,
                  'data[fw_Attach][function_description]': '',
                  'data[fw_Attach][result_of_modified]': '',
                  'data[fw_Attach][note][]': "引用库存软件",
                  'data[change_task_status]': 'on'
                  }
        if disable_origin_sw:
            sw_list = self.get_enable_software_list(ocs_number=ocs_num, exclude_bin=False)
            if sw_list:
                for sw in sw_list:
                    print('sw',sw)
                    fields.update({'data[fw_Attach][fileList][' + sw.attachment_id + ']': 'on'})
                    print('fields',fields)
            print('sw_list',sw_list)
        print('final_fields',fields)
        ret = PyocsRequest().pyocs_request_post(self._upload_fw_url_prefix + str(ocs_num), data=fields)
        return ret.status_code == 200

    #sw_list为空的时候也有返回值
    def set_software_confirm(self, sw_name, order_info, confirm_type=4):
        """
        # 根据sw_name和order_info，确认软件
        """
        ocs_number = self.get_ocs_number_from_sw(sw_name, order_info)
        print("my set_software_confirm ocs_number",ocs_number)
        if ocs_number is None:
            pass
        else:
            sw_list = self.get_enable_software_list(ocs_number=ocs_number)

            if sw_list != None:
                print("my set_software_confirm sw_list",sw_list[0].name)
                fw_id = sw_list[0].fw_id
                first_comfirm_reason_string = str(ocs_number)+str("            ")+str("客批号")+str("            ")+str(order_info)
                confirm_status = confirm_type  # 1-未确认，3-口头确认，(default)4-邮件确认，5-不需确认
                data = {
                    'task_id': int(ocs_number),
                    'field': "account_firmware_status",
                    'firmware_id': int(fw_id),
                    'data': confirm_status,
                    'confirm_reason': 1,
                    'first_confirm_reason': str(first_comfirm_reason_string),
                    'second_confirm_reason': str(sw_name),
                }
                ret = PyocsRequest().pyocs_request_post(self._ocs_confirm_url, data=data)
                confirm_result = ret.text
                self._logger.info(confirm_result)
                if int(confirm_result) > 0:
                    return True
                else:
                    self._logger.info("confirm failed")
            else:
                return False
    #To Do:在这里添加判断订单类型，排除掉意向订单.
    def set_software_confirm_by_ocs_number(self, ocs_number, sw_name, order_info, confirm_type=4):
        """
        # 根据sw_name和ocs号（指定订单了），确认软件
        """
        if ocs_number is None:
            pass
        else:
            sw_list = self.get_enable_software_list(ocs_number=ocs_number)
            if sw_list != None:
                fw_id = sw_list[0].fw_id
                first_comfirm_reason_string = str(ocs_number)+str("            ")+str("客批号")+str("            ")+str(order_info)
                confirm_status = confirm_type  # 1-未确认，3-口头确认，(default)4-邮件确认，5-不需确认
                data = {
                    'task_id': int(ocs_number),
                    'field': "account_firmware_status",
                    'firmware_id': int(fw_id),
                    'data': confirm_status,
                    'confirm_reason': 1,
                    'first_confirm_reason': str(first_comfirm_reason_string),
                    'second_confirm_reason': str(sw_name),
                }
                ret = PyocsRequest().pyocs_request_post(self._ocs_confirm_url, data=data)
                confirm_result = ret.text
                self._logger.info(confirm_result)
                if int(confirm_result) > 0:
                    return True
                else:
                    self._logger.info("confirm failed")
            else:
                return False

#1、引用的时候把指定软件名启用，引用完再停用回来
#2、把所有指向启用启用函数的条件改为指向指定软件名
#3、只通过sw_name查找fw_id来获取这些信息
    def reuse_old_sw_from_src_to_dst_by_fw_id(self, old_sw_name, dst_ocs, workspace):
        """
        引用库存软件，如果此单上已经引用了同名的库存软件，则不做引用
        :param src_ocs:
        :param dst_ocs:
        :return:
        """
        dst_ocs_sw_name_list = []
        reuse_sw_name_list = []
        sw_reuse_burn_type = {'在线烧录': '1', '离线烧录': '2'}
        print("old_sw_name",old_sw_name)
        print("type_old_sw_name",type(old_sw_name))
        if old_sw_name=="" or old_sw_name==" " or old_sw_name=='None':
            #raise NoSoftwareError("源订单上无可用软件")
            return False
        #src_ocs_link = self._ocs_link_Prefix + src_ocs
        #response = PyocsRequest().pyocs_request_get(src_ocs_link)
        #html = etree.HTML(response.text)

        dst_ocs_sw_list = self.get_enable_software_list(ocs_number=dst_ocs, exclude_bin=False)
        #dst_ocs_link = self._ocs_link_Prefix + dst_ocs
        #response = PyocsRequest().pyocs_request_get(dst_ocs_link)
        #html = etree.HTML(response.text)
        #获取的是所有的附件，不是所有的软件
        #dst_ocs_sw_list = self.get_all_software_list_from_html(html)
        #print("dst_ocs_sw_list",dst_ocs_sw_list)

        if dst_ocs_sw_list:
            for dst_ocs_sw in dst_ocs_sw_list:
                # OCS系统的逆天bug，可能会出现两单相同软件，但是其中一个字母一边大写一边小写，此处为了这个bug，统一采用大写比较
                dst_ocs_sw_name_list.append(str(dst_ocs_sw.name.upper()))
        #改sw_list为获取指定sw_name
        #sw_list = self.get_enable_software_list(ocs_number=src_ocs, exclude_bin=False)
        print('dst_ocs_sw_name_list',dst_ocs_sw_name_list)
        print('dst_ocs_sw_list',dst_ocs_sw_list)        
        old_sw_dict=self.find_old_sw_id_by_name(old_sw_name)
        print("old_sw_dict",old_sw_dict)
        #sw_list = old_sw_dict.values()
        sw_list = old_sw_dict.items()
        sw_name_list = old_sw_dict.values()
        sw_fw_id_list = old_sw_dict.keys()
        print("sw_list",sw_list)
        #sw_dict = self.get_enable_software_info_from_html(html, exclude_bin=False)
        if not sw_list:
            #raise NoSoftwareError("源订单上无可用软件")
            return False
        index=0
        for key_fw_id,value_sw_name in sw_list:
            index =index + 1
            print("index",index)
            if  str(value_sw_name).upper() not in dst_ocs_sw_name_list:
                # 如果库存软件包含烧录位号信息，同步引用该信息；否则，根据dst订单自带的烧录位号信息设置
                    #根据dst订单自带的烧录位号信息设置，为什么是src_ocs,可不可以去掉src_ocs改为dst_ocs
                ddr_info_str = PyocsDemand(dst_ocs).get_ddr_info()
                ddr_info_dict = eval(ddr_info_str)
                flash_list1 = ['EMMC FLASH', 'NAND FLASH']
                flash_list2 = ['NOR FLASH']
                burn_place_hold_nums = ddr_info_str.count('refDec')
                burn_place_hold_itemNo = ddr_info_dict['categoryDescp']
                if 1 == burn_place_hold_nums and ddr_info_dict['refDec'] is not None:
                    burn_place_hold = ddr_info_dict['refDec']
                    if burn_place_hold_itemNo in flash_list1:
                        burn_type = self.sw_burn_type["在线烧录"]
                    elif burn_place_hold_itemNo in flash_list2:
                        burn_type = self.sw_burn_type["离线烧录"]
                    if 'EMMCBIN' in str(value_sw_name):
                        burn_type = self.sw_burn_type["离线烧录"]
                else:
                    burn_place_hold = ' '
                    burn_type = ' '
                self.upload_old_sw_by_id(ocs_num=dst_ocs, old_sw_id=key_fw_id, burn_place_hold=burn_place_hold, burn_type=burn_type, disable_origin_sw=(index == 1))
                reuse_sw_name_list.append(value_sw_name)
                print('reuse_sw_name_list',reuse_sw_name_list)
        if not reuse_sw_name_list:
            return True
        while reuse_sw_name_list[-1] \
                not in self.get_enable_software_list_from_html(PyocsDemand(ocs_number=dst_ocs).get_ocs_html()):
            #这个continue是在什么循环里的
            #目的订单没有软件的时候是会编译报错的，argument of type 'NoneType' is not iterable
            continue
        index=0

    def new_reuse_old_sw_from_src_to_dst(self, src_ocs, dst_ocs, workspace):
        """
        引用库存软件，如果此单上已经引用了同名的库存软件，则不做引用
        :param src_ocs:
        :param dst_ocs:
        :return:
        """
        reuse_sw_name_list = faas_api.copy_software(src=src_ocs, dst=dst_ocs)
        if not reuse_sw_name_list:
            return True
        pd = PyocsDemand(ocs_number=dst_ocs)
        customer = pd.get_ocs_customer()
        engineer = pd.get_ocs_software_engineer()
        project = pd.get_ocs_project_name()

    #已经处理过的评审单号不再处理
    def write_order_have_deal(self,order_info):
        # Workbook属于创建才用
        # 打开Excel文件读取数据
        
        #用vscode运行时，用这个
        #path = dirPath+'/customers/customer_jinpin/atuo_do_daice_sw/已处理过的评审单号.xlsx'
        #用windows自动运行时用这个路径，并需要把pyocs放到python安装的lib\site-packages
        path = self.dirPath+'/已处理过的评审单号.xlsx'

        writebook = load_workbook(path)
        # 获取读入的文件的sheet
        sheet = writebook.active
        max_rows_new = sheet.max_row
        row_index = 1
        print("max_rows_new:",max_rows_new)
        while row_index <= max_rows_new:
            b = sheet.cell(row=row_index, column=1)
            row_index = row_index+1
            if None != b.value:
                if order_info == str(b.value):
                    return False

        row = [order_info]
        print(row)
        sheet.append(row)
        writebook.save(path)
        return True
    def get_enable_software_list(self, ocs_number, exclude_bin=True, exclude_lock_sw=True):
        """
        获取订单中启用状态下的软件列表，以Software对象列表返回
        """
        sw_list = list()
        ocs_link = self._ocs_link_Prefix + ocs_number
        response = PyocsRequest().pyocs_request_get(ocs_link)
        html = etree.HTML(response.text)
        sw_name_list = self.get_enable_software_list_from_html(html, exclude_lock_sw)
        if sw_name_list is None:
            return None
        for sw_name in sw_name_list:
            sw = self.Software()
            if exclude_bin and self.is_burn_bin(sw_name=sw_name):
                continue
            if exclude_bin and not self.is_cvte_spec_sw(sw_name=sw_name):
                continue
            sw.name = sw_name
            sw.attachment_id = self.get_enable_software_attachment_id_from_html_by_sw_name(html, sw_name)
            sw.fw_id = self.get_enable_software_fw_id_from_html_by_sw_name(html, sw_name)
            sw_list.append(sw)
        self._logger.info("软件名和attchment id的字典：" + str(sw_list))
        if sw_list:
            return sw_list
        else:
            return None

    def get_enable_software_list_from_html(self, html, exclude_lock_sw = True):
        """获取订单上的启用状态下的软件列表
        Args:
            html: 某个ocs订单的html页面
        Returns:
            启用状态下的软件列表--->所有软件
        """
        enable_software_list = []
        # enable_software_name_xpath = '//strong[text()="同时停用以下软件包"]/../../..//input[@type="checkbox"]/..//label/text()'  # '//a[@title="停用软件"]/@file_name'
        enable_software_name_xpath = \
            '//a[contains(@href, "/tv/Attachments/download_attachment")]/text()'
        enable_software_name_list = html.xpath(enable_software_name_xpath)
        self._logger.info("启用状态的软件:" + str(enable_software_name_list))
        if not exclude_lock_sw and enable_software_name_list:
            return enable_software_name_list

        for sw_name in enable_software_name_list:
            if not self.is_sw_locked(sw_name=sw_name, html=html):
                enable_software_list.append(sw_name)

        if enable_software_list:
            return enable_software_list
        else:
            return None

    def get_enable_software_list_name_from_html(self, html, exclude_lock_sw = True):
        """获取订单上的启用状态下的软件列表
        Args:
            html: 某个ocs订单的html页面
        Returns:
            启用状态下的软件列表--->所有软件
        """
        enable_software_list = []
        # enable_software_name_xpath = '//strong[text()="同时停用以下软件包"]/../../..//input[@type="checkbox"]/..//label/text()'  # '//a[@title="停用软件"]/@file_name'
        enable_software_name_xpath = \
            '//a[contains(@href, "/tv/Attachments/download_attachment")]/text()'
        enable_software_name_list = html.xpath(enable_software_name_xpath)
        self._logger.info("启用状态的软件:" + str(enable_software_name_list))
        if not exclude_lock_sw and enable_software_name_list:
            return enable_software_name_list

        for sw_name in enable_software_name_list:
            if not self.is_sw_locked(sw_name=sw_name, html=html):
                enable_software_list.append(sw_name)

        if enable_software_list:
            return enable_software_list
        else:
            return None


    def get_all_software_list_from_html(self, html, exclude_bin=True):
        """获取订单上的所有软件
        Args:
            html: 某个ocs订单的html页面
        Returns:
            所有软件，以列表返回
        """
        sw_list = list()
        sw_name_list = html.xpath('//div[@id="file_names"]/p/text()')
        print("sw_name_list",sw_name_list)
        for sw_name in sw_name_list:
            sw = self.Software()
            if exclude_bin and self.is_burn_bin(sw_name=sw_name):
                continue
            if exclude_bin and not self.is_cvte_spec_sw(sw_name=sw_name):
                continue
            sw.name = sw_name
            sw.attachment_id = self.get_enable_software_attachment_id_from_html_by_sw_name(html, sw_name)
            sw.fw_id = self.get_enable_software_fw_id_from_html_by_sw_name(html, sw_name)
            sw_list.append(sw)
        self._logger.info("软件名和attchment id的字典：" + str(sw_list))    
        if sw_list:
            print("get_all_software_list_from_html sw_list",sw_list)
            return sw_list
        else:
            return None                            

    def get_all_software_list_name_from_html(self, html, exclude_bin=True):
        """获取订单上的所有软件
        Args:
            html: 某个ocs订单的html页面
        Returns:
            所有软件，以列表返回
        """
        sw_list = list()
        sw_name_list = html.xpath('//div[@id="file_names"]/p/text()')
        print("sw_name_list",sw_name_list)  
        if sw_name_list:
            print("get_all_software_list_from_html sw_list",sw_name_list)
            return sw_name_list
        else:
            return None   

    def get_all_file_list_from_html(self, ocs, exclude_lock_sw = True):
        dst_ocs_link = self._ocs_link_Prefix + ocs
        print("dst_ocs_link",dst_ocs_link)
        response = PyocsRequest().pyocs_request_get(dst_ocs_link)
        html = etree.HTML(response.text)
        #获取的是所有的附件，不是所有的软件
        dst_ocs_sw_list = self.get_all_software_list_from_html(self, html)
        print("dst_ocs_sw_list",dst_ocs_sw_list)
        return dst_ocs_sw_list            