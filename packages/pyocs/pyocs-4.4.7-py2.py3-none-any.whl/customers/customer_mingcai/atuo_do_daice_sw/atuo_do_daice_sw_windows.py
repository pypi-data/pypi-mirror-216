import json
import requests
from requests.exceptions import RequestException
import re
import time
import random
import os
import logging
from lxml import etree
import sys
print(sys.path)
sys.path.append(".")
import time
import jenkins
from openpyxl import load_workbook
from pyocs.pyocs_software import PyocsSoftware
from pyocs.pyocs_demand import PyocsDemand
from pyocs.pyocs_request import PyocsRequest
from pyocs import pyocs_login
from pyocs.pyocs_list import PyocsList
from pyocs import pyocs_login
from pyocs.pyocs_request import PyocsRequest
import logging
from urllib.parse import urlencode
from lxml import etree
from colorama import Fore, Back, Style
from pyocs.pyocs_jenkins import PyocsJenkins
import platform

new_ocs_url = "http://ocs.gz.cvte.cn/tv/Tasks/index/ContractTask//SearchId:2091062/range:all"
find_old_keliaohao_ocs_url = "http://ocs.gz.cvte.cn/tv/Tasks/index/range:all/"
auto_build_url = "http://tvci.gz.cvte.cn/job/yangpingbu/job/Sample_Compile/build"
reuse_stock_software_url = "http://tvci.gz.cvte.cn/job/Utility_Tools/job/reuse_stock_software/build"
customer_name ="金品"
customer_order = "01-202005-132"
customer_liaohao="R41-63S2T2B00000A1C"
member = "林祥纳"
#member = "陈潮雄"
ocs_status="待录入需求"
#list_xpath = '//td[@class="Task_col_id"]|//td[@class="Task_col_sw_user_id"]|//td[@class="Task_col_status"]|//td[@class="Task_col_subject"]/a|//td[@class="Task_col_rd_dept_id"]'
list_xpath = '//td[@class="Task_col_id"]|//td[@class="Task_col_sw_user_id"]|//td[@class="Task_col_status"]|//td[@class="Task_col_subject"]/a|//td[@class="Task_col_rd_dept_id"]|//td[@class="Contract_col_account_prod_name"]'


#dirPath = os.getcwd()+'\\customers\\customer_jinpin\\atuo_do_daice_sw'
dirPath = os.getcwd()
print("dirPath",dirPath)
class atuo_do_daice_sw:

    _logger = logging.getLogger(__name__)

    def __init__(self):
        self._logger.setLevel(level=logging.WARNING)  # 控制打印级别

    def get_new_order_searchid(self,advanced_search,condition):
        #根据摘要里的条件获取search_id
        searchid = PyocsSoftware().get_searchid_from_advanced_search(advanced_search, condition)
        return searchid

    def write_ocs_have_support_auto(self,ocs_number):
        # Workbook属于创建才用
        # 打开Excel文件读取数据
        
        #用vscode运行时，用这个
        #path = dirPath+'\\customers\\customer_jinpin\\atuo_do_daice_sw\\已提交自动编译.xlsx'
        #用windows自动运行时用这个路径，并需要把pyocs放到python安装的lib\site-packages
        path = dirPath+'\\已提交自动编译.xlsx'

        writebook = load_workbook(path)
        # 获取读入的文件的sheet
        sheet = writebook.active
        max_rows_new = sheet.max_row
        row_index = 1
        print("max_rows_new:",max_rows_new)
        while row_index <= max_rows_new:
            b = sheet.cell(row=row_index, column=1)
            row_index = row_index+1
            if None != b.value:
                if ocs_number == b.value:
                    return False

        row = [ocs_number]
        print(row)
        sheet.append(row)
        writebook.save(path)
        return True
    
    def write_ocs_have_submit_reuse(self,ocs_number):
        # Workbook属于创建才用
       
        #用vscode运行时，用这个
        #path = dirPath+'\\customers\\customer_jinpin\\atuo_do_daice_sw\\已提交引用.xlsx'
        #用windows自动运行时用这个路径，并需要把pyocs放到python安装的lib\site-packages
        path = dirPath+'\\已提交引用.xlsx'
        writebook = load_workbook(path)
        # 获取读入的文件的sheet
        sheet = writebook.active
        max_rows_new = sheet.max_row
        row_index = 1
        print("max_rows_new:",max_rows_new)
        while row_index <= max_rows_new:
            b = sheet.cell(row=row_index, column=1)
            row_index = row_index+1
            if None != b.value:
                if ocs_number == b.value:
                    return False

        row = [ocs_number]
        print(row)
        sheet.append(row)
        writebook.save(path)
        return True

    def change_list_to_2D_list(self,new_ocs_list):
        new_ocs_order_list = list()
        i = 0
        parse_list_width = 6 
        for j in range(0,len(new_ocs_list),parse_list_width):
            new_ocs_order_list.append(new_ocs_list[j:j+parse_list_width])
            i=i+1
        return i,new_ocs_order_list

    def parse_string_from_ocs_info(self,ocs_subjest_info,string_pattern):
        pattern = re.compile(string_pattern, re.S)
        print(pattern)
        items = re.findall(pattern, ocs_subjest_info)
        print(items)
        return items       

    def submit_auto_compile(self,ocs_number_list,nums):
        jenkins_server = jenkins.Jenkins
        account = pyocs_login.PyocsLogin().get_account_from_json_file()
        jenkins_server = jenkins.Jenkins('http://tvci.gz.cvte.cn/', username=account['Username'], password=account['Password'],
                                    timeout=30)
        
        check_if_have_submit_auto_flag = False
        for i in range(0,nums):
            ocs_number = ocs_number_list[i][0]
            if "TCL" not in ocs_number_list[i][3] and "[虚拟]" not in ocs_number_list[i][3] :
                auto_make_bin_flag = do_daice_task.check_if_auto_make_bin(ocs_number_list[i][4])
                surport_auto_compile_flag = do_daice_task.check_if_project_support_auto_compile(ocs_number_list[i][4])
                check_if_have_submit_auto_flag =self.write_ocs_have_support_auto(ocs_number)
                data = {
                    'OCS_NUMBER': int(ocs_number),
                    'AUTO_MAKE_BIN': str(auto_make_bin_flag),
                    'COMPILE_SOFTWARE': 'True'
                }
                #print(data)
                #r_jenkins=requests.post(auto_build_url, data=data)
                #if (check_if_have_submit_auto_flag is True) and (surport_auto_compile_flag is True):
                #surport_auto_compile_flag获取的是string类型
                if (check_if_have_submit_auto_flag is True) and (str(surport_auto_compile_flag) == "True"):
                #if  (str(surport_auto_compile_flag) == "True"):
                    print("submit_auto_compile",ocs_number)
                    r_jenkins=jenkins_server.build_job("yangpingbu/Sample_Compile" ,data)     

    def submit_atuo_reuse(self,src_ocs_number,dst_ocs_number):
        jenkins_server = jenkins.Jenkins
        account = pyocs_login.PyocsLogin().get_account_from_json_file()
        jenkins_server = jenkins.Jenkins('http://tvci.gz.cvte.cn/', username=account['Username'], password=account['Password'],
                                    timeout=30)
        print(jenkins_server)
        data = {
            'SRC_OCS_NUMBER': int(src_ocs_number),
            'DST_OCS_NUMBER': int(dst_ocs_number),
        }
        #print(data)
        r_jenkins=jenkins_server.build_job("Utility_Tools/reuse_stock_software" ,data)

    def check_if_auto_make_bin(self,str_fangan):
        
        #window下运行时，用这个
        _support_auto_make_bin_file = dirPath+'\\support_auto_make_bin.json'
        try:
            logging.info(_support_auto_make_bin_file)
            with open(_support_auto_make_bin_file, 'r') as load_file:
                tmp = load_file.read()
                support_auto_make_bin_dict = json.loads(json.dumps(eval(tmp)))
                if  str(str_fangan) in support_auto_make_bin_dict:
                    return support_auto_make_bin_dict[str_fangan]
                else:
                    return False                   
                #return support_auto_make_bin_dict
        except FileNotFoundError:
            logging.error("请查找运行路径是否有保存支持自动做bin方案的json文件")
            #先默认支持做bin.支持做bin没什么坏处
            return True

    def check_if_project_support_auto_compile(self,str_fangan):
        
        #window下运行时，用这个
        #getcwd()获取行时所在目录，对运行路径要求苛刻
        _support_auto_make_bin_file = dirPath+'\\support_auto_compile.json'
        try:
            logging.info(_support_auto_make_bin_file)
            with open(_support_auto_make_bin_file, 'r') as load_file:
                tmp = load_file.read()
                support_auto_compile_dict = json.loads(json.dumps(eval(tmp)))
                if str(str_fangan) in support_auto_compile_dict:
                    return support_auto_compile_dict[str_fangan]
                else:
                    return False
                #return support_auto_make_bin_dict
        except FileNotFoundError:
            logging.error("请查找运行路径是否有保存支持自动编译方案的json文件")
            #先默认支持自动编译.支持自动编译稍微浪费了点服务器资源
            return True

    def parse_lock_software_and_reason(self,ocs_list):
        flag = True
        return

    #还需要加功能，判断订单返回的可用软件列表不为空才应用
    def check_if_reuse_old(self,new_ocs_info_list,count_nums):
        #如果没有单要处理，old_ocs_number没有赋值会导致最后return错误
        old_ocs_number = 0
        for i in range(0,count_nums):
            dst_ocs_number = str(new_ocs_info_list[i][0])
            print("dst_ocs_number",dst_ocs_number)
            pd_dst = PyocsDemand(ocs_number=dst_ocs_number)
            print("产品型号",pd_dst.get_product_name())
            dst_product_name = pd_dst.get_product_name()
            dst_pwm = pd_dst.get_pwm()
            dst_port_info = pd_dst.get_port_info()
            #每个新的单重新赋值，归零可以引用的旧软件OCS,以免用了旧的值
            old_ocs_number = 0
            #现在是不重复reuse.但是后面做获取禁用软件信息申请解禁。还要可以继续reuse
            string_pattern_liaohao='\u5ba2\u6599\u53f7(.*?)，'
            string_pattern_model_size='E(\d\d)'
            ocs_subjest_info  = u"[B.NEW][金品]TP.ATM30.PB818C配ST3751A07-2（中国，SY202005150001，客料号R41-920CN230C0AWA1C，客批号TCL海外电子01-202005-101）"
            ocs_subjest_info  = new_ocs_info_list[i][3]
            if '[虚拟]' in ocs_subjest_info:
                continue
            string_model_size  = do_daice_task.parse_string_from_ocs_info(new_ocs_info_list[i][5],string_pattern_model_size)
            string_liaohao=do_daice_task.parse_string_from_ocs_info(ocs_subjest_info,string_pattern_liaohao)
            #一开始string_liaohao是list类型，要剃干净，没有一点多余的字符才行
            string_liaohao =str(string_liaohao)
            string_liaohao = string_liaohao[2:-2].split("-", 2)[1]
            print("after split",string_liaohao)
            liaohao_advanced_search = {
                "0":{"search_field_name":"Task.subject","search_field_id":"554","search_field_type":"5","search_field_rel_obj":"null","search_opr":"TDD_OPER_INC","input1":string_liaohao,"input2":"null","offset":"null"}
                }
            new_order_condition = ""
            search_id_liaohao = do_daice_task.get_new_order_searchid(liaohao_advanced_search,new_order_condition)
            #print(search_id)
            xpath_result_liaohao = PyocsList().get_single_data_list(list_xpath,search_id_liaohao)
            temp_list=do_daice_task.change_list_to_2D_list(xpath_result_liaohao[1])
            ocs_list_liaohao = temp_list[1]
            count_nums = temp_list[0]
            #print(temp_list)
            #print(count_nums)
            #wifi模块的软件不做处理
            #if(new_ocs_info_list[i][4] != "MT7601") and (new_ocs_info_list[i][4] != "6M60") :
            if(new_ocs_info_list[i][4] != "MT7601") :
                for j in range(count_nums-1,-1,-1):
                    #print("j:",j)
                    #现在是找ocs上有未被禁用软件的订单
                    if ocs_list_liaohao[j][2] != "软件审核不通过" and ocs_list_liaohao[j][2] != "软件测试不通过" and ocs_list_liaohao[j][2] != "待录入需求":
                        if do_daice_task.get_enable_software_can_be_reuse_with_ocs(ocs_list_liaohao[j][0]) == None:
                            continue
                        else:
                            print("ocs_list_liaohao[j][0]",ocs_list_liaohao[j][0])
                            #print("get_enable_software_reuse_ocs_with_ocs",PyocsSoftware().get_enable_software_reuse_ocs_with_ocs(ocs_list_liaohao[j][0]))
                            old_ocs_string_model_size  = do_daice_task.parse_string_from_ocs_info(ocs_list_liaohao[j][5],string_pattern_model_size)
                            #print(old_ocs_string_model_size)
                            old_backup_ocs_number = str(ocs_list_liaohao[j][0])
                            pd_old = PyocsDemand(ocs_number=old_backup_ocs_number)
                            print("产品型号",pd_old.get_product_name())
                            old_product_name = pd_old.get_product_name()
                            old_pwm = pd_old.get_pwm()
                            old_port_info = pd_old.get_port_info()
                            #if  old_ocs_string_model_size == string_model_size:
                            if  (dst_product_name == old_product_name) and (dst_pwm == old_pwm) and (dst_port_info == old_port_info):
                                old_ocs_number= ocs_list_liaohao[j][0]
                                break
            print("old_ocs_number",old_ocs_number)
            if old_ocs_number!=0:
                do_daice_task.submit_atuo_reuse(old_ocs_number,dst_ocs_number)
        return old_ocs_number

    def get_enable_software_list_with_ocs(self, ocs, exclude_bin=True):
        """获取订单上的启用状态下的软件以及对应的信息字典
        Args:
            html: 某个ocs订单的html页面
        Returns:
            启用状态下的非烧录 bin 或者不符合 CVTE 规范的软件以及对应的信息字典
        """
        active_software_info_list = []
        _ocs_base_link = 'https://ocs.gz.cvte.cn'
        _ocs_link_Prefix = _ocs_base_link + '/tv/Tasks/view/'
        src_ocs_link = _ocs_link_Prefix + str(ocs)
        response = PyocsRequest().pyocs_request_get(src_ocs_link)
        html = etree.HTML(response.text)
        enable_software_name_xpath = \
            '//a[contains(@href, "/tv/Attachments/download_attachment")]/text()'
        enable_software_name_list = html.xpath(enable_software_name_xpath)
        return enable_software_name_list


    def get_enable_software_can_be_reuse_with_ocs(self, ocs, exclude_bin=True):
        """获取订单上的启用状态下的软件以及对应的信息字典
        Args:
            html: 某个ocs订单的html页面
        Returns:
            启用状态下的非烧录 bin 或者不符合 CVTE 规范的软件以及对应的信息字典
        """
        active_software_info_list = []
        _ocs_base_link = 'https://ocs.gz.cvte.cn'
        _ocs_link_Prefix = _ocs_base_link + '/tv/Tasks/view/'
        src_ocs_link = _ocs_link_Prefix + str(ocs)
        response = PyocsRequest().pyocs_request_get(src_ocs_link)
        html = etree.HTML(response.text)
        enable_software_name_xpath = \
            '//a[contains(@href, "/tv/Attachments/download_attachment")]/text()'
        enable_software_name_list = html.xpath(enable_software_name_xpath)
        for sw_name in enable_software_name_list:
            if not PyocsSoftware().is_sw_locked(sw_name=sw_name, html=html):
                active_software_info_list.append(sw_name)
        print("active_software_info_list",active_software_info_list)
        if active_software_info_list:
            enable_reuse_software_info = active_software_info_list[0]
            enable_reuse_ocs_list = re.findall(r'\d+', enable_reuse_software_info)
            return active_software_info_list
        else:
            return None


class Logger(object):
    def __init__(self, filename='default.log', stream=sys.stdout, encoding='utf-8'):
        self.terminal = stream
        self.log = open(filename, 'a')

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        pass

sys.stdout = Logger(stream=sys.stdout)
sys.stdout = Logger(stream=sys.stderr)

sys_time=time.strftime('time:%Y-%m-%d-%H-%M-%S', time.localtime(time.time()))
print(sys_time)

do_daice_task = atuo_do_daice_sw()
pyocs_job = PyocsSoftware()
#new_order_advanced_search = {
#    "0": {"search_field_name":"Task.sw_user_id","search_field_id":"555","search_field_type":"19","search_field_rel_obj":"Users","search_opr":"TDD_OPER_INC","input1":member,"input2":"null","offset":"null"},
#    "1": {"search_field_name":"Task.status","search_field_id":"584","search_field_type":"19","search_field_rel_obj":"Enums","search_opr":"TDD_OPER_EQUAL","input1":ocs_status,"input2":"null","offset":"null"}
#    }
#金品 且待录入需求的订单
new_order_advanced_search = {
    "0":{"search_field_name":"Task.account_id","search_field_id":"560","search_field_type":"19","search_field_rel_obj":"Accounts","search_opr":"TDD_OPER_INC","input1":"\u91d1\u54c1","input2":"null","offset":"null"},
    "1":{"search_field_name":"Task.status","search_field_id":"584","search_field_type":"19","search_field_rel_obj":"Enums","search_opr":"TDD_OPER_INC","input1":ocs_status,"input2":"null","offset":"null"}
    }
new_order_condition = "(1 and 2)"
submit_auto_compile_flag=0


search_id_new_order = do_daice_task.get_new_order_searchid(new_order_advanced_search,new_order_condition)
print(search_id_new_order)

xpath_result = PyocsList().get_single_data_list(list_xpath,search_id_new_order)
#print(xpath_result)
new_ocs_list = xpath_result[1]
print(new_ocs_list)
#count_ocs_numbers = PyocsList().get_ocs_id_list(search_id)
temp_list=do_daice_task.change_list_to_2D_list(new_ocs_list)
new_ocs_info_list = temp_list[1]
count_nums = temp_list[0]
#print(count_nums)
#print(new_ocs_info_list[0])
do_daice_task.check_if_reuse_old(new_ocs_info_list,count_nums)

submit_auto_compile_flag = 1


#reuse成功后，应该就没有待录入需求目录了
#大致等个10分钟，等reuse完。该方法不是很好。应该是直接获取reuse成功状态
time.sleep(600)
search_id_new_order = do_daice_task.get_new_order_searchid(new_order_advanced_search,new_order_condition)
print(search_id_new_order)

xpath_result = PyocsList().get_single_data_list(list_xpath,search_id_new_order)
#print(xpath_result)
new_ocs_list = xpath_result[1]
print(new_ocs_list)
#count_ocs_numbers = PyocsList().get_ocs_id_list(search_id)
temp_list=do_daice_task.change_list_to_2D_list(new_ocs_list)
new_ocs_info_list = temp_list[1]
count_nums = temp_list[0]
do_daice_task.submit_auto_compile(new_ocs_info_list,count_nums)




