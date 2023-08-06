import json
import requests
from requests.exceptions import RequestException
import re
import time
import random
import os
import logging
from lxml import etree
import sys
print(sys.path)
sys.path.append(".")
import time
import jenkins
from openpyxl import load_workbook
from pyocs.pyocs_software import PyocsSoftware
from pyocs.pyocs_demand import PyocsDemand
from pyocs.pyocs_request import PyocsRequest
from pyocs import pyocs_login
from pyocs.pyocs_list import PyocsList
from pyocs import pyocs_login
from pyocs.pyocs_request import PyocsRequest
import logging
from urllib.parse import urlencode
from lxml import etree
from colorama import Fore, Back, Style
from pyocs.pyocs_jenkins import PyocsJenkins

new_ocs_url = "http://ocs.gz.cvte.cn/tv/Tasks/index/ContractTask//SearchId:2091062/range:all"
find_old_keliaohao_ocs_url = "http://ocs.gz.cvte.cn/tv/Tasks/index/range:all/"
auto_build_url = "http://tvci.gz.cvte.cn/job/yangpingbu/job/Sample_Compile/build"
reuse_stock_software_url = "http://tvci.gz.cvte.cn/job/Utility_Tools/job/reuse_stock_software/build"
customer_name ="金品"
customer_order = "01-202005-132"
customer_liaohao="R41-63S2T2B00000A1C"
member = "林祥纳"
#member = "陈潮雄"
ocs_status="待录入需求"
#list_xpath = '//td[@class="Task_col_id"]|//td[@class="Task_col_sw_user_id"]|//td[@class="Task_col_status"]|//td[@class="Task_col_subject"]/a|//td[@class="Task_col_rd_dept_id"]'
list_xpath = '//td[@class="Task_col_id"]|//td[@class="Task_col_sw_user_id"]|//td[@class="Task_col_status"]|//td[@class="Task_col_subject"]/a|//td[@class="Task_col_rd_dept_id"]|//td[@class="Contract_col_account_prod_name"]'

class atuo_do_daice_sw:


    def get_new_order_searchid(self,advanced_search,condition):
        #根据摘要里的条件获取search_id
        searchid = PyocsSoftware().get_searchid_from_advanced_search(advanced_search, condition)
        return searchid

    def write_ocs_have_support_auto(self,ocs_number):
        # Workbook属于创建才用
        # 打开Excel文件读取数据
        dirPath = os.getcwd()
        #用vscode运行时，用这个
        path = dirPath+'\\customers\\customer_jinpin\\atuo_do_daice_sw\\已提交自动编译.xlsx'
        #用windows自动运行时用这个路径，并需要把pyocs放到python安装的lib\site-packages
        #path = dirPath+'\\已提交自动编译.xlsx'

        writebook = load_workbook(path)
        # 获取读入的文件的sheet
        sheet = writebook.active
        max_rows_new = sheet.max_row
        row_index = 1
        print("max_rows_new:",max_rows_new)
        while row_index <= max_rows_new:
            b = sheet.cell(row=row_index, column=1)
            row_index = row_index+1
            if None != b.value:
                if ocs_number == b.value:
                    return False

        row = [ocs_number]
        print(row)
        sheet.append(row)
        writebook.save(path)
        return True
    
    def write_ocs_have_submit_reuse(self,ocs_number):
        # Workbook属于创建才用
        # 打开Excel文件读取数据
        dirPath = os.getcwd()
        #用vscode运行时，用这个
        path = dirPath+'\\customers\\customer_jinpin\\atuo_do_daice_sw\\已提交引用.xlsx'
        #用windows自动运行时用这个路径，并需要把pyocs放到python安装的lib\site-packages
        #path = dirPath+'\\已提交引用.xlsx'
        writebook = load_workbook(path)
        # 获取读入的文件的sheet
        sheet = writebook.active
        max_rows_new = sheet.max_row
        row_index = 1
        print("max_rows_new:",max_rows_new)
        while row_index <= max_rows_new:
            b = sheet.cell(row=row_index, column=1)
            row_index = row_index+1
            if None != b.value:
                if ocs_number == b.value:
                    return False

        row = [ocs_number]
        print(row)
        sheet.append(row)
        writebook.save(path)
        return True

    def change_list_to_2D_list(self,new_ocs_list):
        new_ocs_order_list = list()
        i = 0
        parse_list_width = 6 
        for j in range(0,len(new_ocs_list),parse_list_width):
            new_ocs_order_list.append(new_ocs_list[j:j+parse_list_width])
            i=i+1
        return i,new_ocs_order_list

    def parse_string_from_ocs_info(self,ocs_subjest_info,string_pattern):
        pattern = re.compile(string_pattern, re.S)
        print(pattern)
        items = re.findall(pattern, ocs_subjest_info)
        print(items)
        return items

    def submit_auto_compile(self,ocs_number_list,nums,submit_auto_compile_flag):
        jenkins_server = jenkins.Jenkins
        account = pyocs_login.PyocsLogin().get_account_from_json_file()
        jenkins_server = jenkins.Jenkins('http://tvci.gz.cvte.cn/', username=account['Username'], password=account['Password'],
                                    timeout=30)
        print(jenkins_server)
        check_if_have_support_auto_flag = False
        for i in range(0,nums):
            ocs_number = ocs_number_list[i][0]
            check_if_have_support_auto_flag =self.write_ocs_have_support_auto(ocs_number)
            data = {
                'OCS_NUMBER': int(ocs_number),
                'AUTO_MAKE_BIN': 'True',
                'COMPILE_SOFTWARE': 'True'
            }
            #print(data)
            #r_jenkins=requests.post(auto_build_url, data=data)
            if check_if_have_support_auto_flag is True:
                r_jenkins=jenkins_server.build_job("yangpingbu/Sample_Compile" ,data)     

    def submit_atuo_reuse(self,src_ocs_number,dst_ocs_number):
        jenkins_server = jenkins.Jenkins
        account = pyocs_login.PyocsLogin().get_account_from_json_file()
        jenkins_server = jenkins.Jenkins('http://tvci.gz.cvte.cn/', username=account['Username'], password=account['Password'],
                                    timeout=30)
        print(jenkins_server)
        data = {
            'SRC_OCS_NUMBER': int(src_ocs_number),
            'DST_OCS_NUMBER': int(dst_ocs_number),
        }
        #print(data)
        r_jenkins=jenkins_server.build_job("Utility_Tools/reuse_stock_software" ,data)

    def check_if_auto_make_bin(self,ocs_list):
        flag = True
        return

    def check_if_project_support_auto_compile(self,ocs_list):
        flag = True
        return
    def parse_lock_software_and_reason(self,ocs_list):
        flag = True
        return
    #还需要加功能，判断订单返回的可用软件列表不为空才应用
    #同一料号还有占空比不一样的情况（较少）
    #找旧软件应该要尽量找已经测试通过的
    def check_if_reuse_old(self,new_ocs_info_list,count_nums):
        for i in range(0,count_nums):
            dst_ocs_number = str(new_ocs_info_list[i][0])
            print("dst_ocs_number",dst_ocs_number)
            old_ocs_number = 0
            #现在是不重复reuse.但是后面做获取禁用软件信息申请解禁。还要可以继续reuse
            string_pattern_liaohao='\u5ba2\u6599\u53f7(.*?)，'
            string_pattern_model_size='E(\d\d)'
            ocs_subjest_info  = u"[B.NEW][金品]TP.ATM30.PB818C配ST3751A07-2（中国，SY202005150001，客料号R41-920CN230C0AWA1C，客批号TCL海外电子01-202005-101）"
            ocs_subjest_info  = new_ocs_info_list[i][3]
            string_model_size  = do_daice_task.parse_string_from_ocs_info(new_ocs_info_list[i][5],string_pattern_model_size)
            string_liaohao=do_daice_task.parse_string_from_ocs_info(ocs_subjest_info,string_pattern_liaohao)
            #一开始string_liaohao是list类型，要剃干净，没有一点多余的字符才行
            string_liaohao =str(string_liaohao)
            string_liaohao = string_liaohao[2:-2]
            liaohao_advanced_search = {
                "0":{"search_field_name":"Task.subject","search_field_id":"554","search_field_type":"5","search_field_rel_obj":"null","search_opr":"TDD_OPER_INC","input1":string_liaohao,"input2":"null","offset":"null"}
                }
            new_order_condition = ""
            search_id_liaohao = do_daice_task.get_new_order_searchid(liaohao_advanced_search,new_order_condition)
            #print(search_id)
            xpath_result_liaohao = PyocsList().get_single_data_list(list_xpath,search_id_liaohao)
            temp_list=do_daice_task.change_list_to_2D_list(xpath_result_liaohao[1])
            ocs_list_liaohao = temp_list[1]
            count_nums = temp_list[0]
            #print(temp_list)
            #print(count_nums)
            for j in range(count_nums-1,-1,-1):
                #print("j:",j)
                #现在只是找这两种状态，其实不是这两种状态的也可以找，只要能找到测试通过且未被禁用的就行
                if ocs_list_liaohao[j][2]=="已完成" or ocs_list_liaohao[j][2]=="待上传软件":
                    old_ocs_string_model_size  = do_daice_task.parse_string_from_ocs_info(ocs_list_liaohao[j][5],string_pattern_model_size)
                    #print(old_ocs_string_model_size)
                    if  old_ocs_string_model_size == string_model_size:
                        old_ocs_number= ocs_list_liaohao[j][0]
                        break
            print("old_ocs_number",old_ocs_number)
            if old_ocs_number!=0:
                do_daice_task.submit_atuo_reuse(old_ocs_number,dst_ocs_number)
        return old_ocs_number

do_daice_task = atuo_do_daice_sw()
pyocs_job = PyocsSoftware()
new_order_advanced_search = {
    "0": {"search_field_name":"Task.sw_user_id","search_field_id":"555","search_field_type":"19","search_field_rel_obj":"Users","search_opr":"TDD_OPER_INC","input1":member,"input2":"null","offset":"null"},
    "1": {"search_field_name":"Task.status","search_field_id":"584","search_field_type":"19","search_field_rel_obj":"Enums","search_opr":"TDD_OPER_EQUAL","input1":ocs_status,"input2":"null","offset":"null"}
    }

new_order_condition = "(1 and 2)"
submit_auto_compile_flag=0


search_id_new_order = do_daice_task.get_new_order_searchid(new_order_advanced_search,new_order_condition)
print(search_id_new_order)

xpath_result = PyocsList().get_single_data_list(list_xpath,search_id_new_order)
#print(xpath_result)
new_ocs_list = xpath_result[1]
print(new_ocs_list)
#count_ocs_numbers = PyocsList().get_ocs_id_list(search_id)
temp_list=do_daice_task.change_list_to_2D_list(new_ocs_list)
new_ocs_info_list = temp_list[1]
count_nums = temp_list[0]
#print(count_nums)
#print(new_ocs_info_list[0])
do_daice_task.check_if_reuse_old(new_ocs_info_list,count_nums)

submit_auto_compile_flag = 1


#reuse成功后，应该就没有待录入需求目录了
#大致等个10分钟，等reuse完。改方法不是很好。应该是直接获取reuse成功状态
#time.sleep(600)
search_id_new_order = do_daice_task.get_new_order_searchid(new_order_advanced_search,new_order_condition)
print(search_id_new_order)

xpath_result = PyocsList().get_single_data_list(list_xpath,search_id_new_order)
#print(xpath_result)
new_ocs_list = xpath_result[1]
print(new_ocs_list)
#count_ocs_numbers = PyocsList().get_ocs_id_list(search_id)
temp_list=do_daice_task.change_list_to_2D_list(new_ocs_list)
new_ocs_info_list = temp_list[1]
count_nums = temp_list[0]
do_daice_task.submit_auto_compile(new_ocs_info_list,count_nums,submit_auto_compile_flag)





