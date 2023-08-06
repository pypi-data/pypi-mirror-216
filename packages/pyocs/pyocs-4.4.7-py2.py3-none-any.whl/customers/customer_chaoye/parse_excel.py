#!/usr/bin/env python3

import openpyxl


class ParseExcel:
    excel_sheet = "订单需求"
    request_sheet = None
    request_rows = 0
    request_columns = 0

    # 对应需求表中的列
    ColumnsStr = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]

    # 存储需求表中某一行的软件需求，每一个Excel cell，由[value,comment]的列表存储需求信息
    RequestInfo = {"KEY0": [],
                   "KEY1": [],
                   "KEY2": [],
                   "KEY3": [],
                   "KEY4": [],
                   "KEY5": [],
                   "KEY6": [],
                   "KEY7": [],
                   "KEY8": [],
                   "KEY9": [],
                   "KEY10": [],
                   "KEY11": []}

    def __init__(self, excel_file):
        workbook = openpyxl.load_workbook(excel_file)
        self.request_sheet = workbook.get_sheet_by_name(self.excel_sheet)
        self.request_rows = self.request_sheet.max_row
        self.request_columns = self.request_sheet.max_column

    def _get_order_request_by_number(self, num):
        row = None
        # +2 是有客户需求表头占据了2行
        column_num = "A" + str(int(num) + 2)
        if self.request_sheet[column_num].value == int(num):
            row = int(num) + 2
            return row
        else:
            print("Your No. is not exist in request table, please check request table!")
            return row

    def _get_request_cell_name_by_row(self, row):
        cell_names = []
        for column in self.ColumnsStr:
            name = column + str(row)
            cell_names.append(name)
        return cell_names

    def _write_value_to_request_info(self, cells):
        index = 0
        for i in range(0, len(cells)):
            cell_name = cells[i]
            key_index = "KEY" + str(index)
            cell_value = self.request_sheet[cell_name].value

            self.RequestInfo[key_index].append(cell_value)

            if not self.request_sheet[cell_name].comment is None:
                cell_comment = self.request_sheet[cell_name].comment.text
                self.RequestInfo[key_index].append(cell_comment)

            index = index + 1

    def get_order_request(self, num):
        row = self._get_order_request_by_number(num)
        if row is None:
            return

        cell_names = self._get_request_cell_name_by_row(row)
        self._write_value_to_request_info(cell_names)
        # print(self.RequestInfo)
        return self.RequestInfo