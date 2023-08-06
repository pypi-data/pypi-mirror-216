from pyocs.pyocs_list import PyocsList
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

class OcsSoftware:
    SearchID = "2618675"
    projectID = "6486"
    OCS_LIST = []
    CUSTOMER_LIST = []
    USER_LIST = []
    DATA_LIST = []
    NUMBER_LIST = []
    PACKAGE_LIST = []
    PRODUCT_LIST = []
    COUNTRY_LIST = []
    filepath = os.path.abspath("save_ocs_data.xlsx")

    def __init__(self):
        self.SearchID = "2258175"
        self.projectID = "56"

    def loadOcsData(self):
        (count, self.OCS_LIST) = PyocsList().get_ocs_id_list(self.SearchID)
        if count == 0:
            print("数据为空")
            exit(0)

        (count, self.CUSTOMER_LIST) = PyocsList().get_ocs_customerId_list(self.SearchID)
        (count, self.USER_LIST) = PyocsList().get_ocs_engineer_list(self.SearchID)
        (count, self.DATA_LIST) = PyocsList().get_ocs_specific_list(xpath_id = "Task_col_plan_end_date",search_id = self.SearchID)
        (count, self.NUMBER_LIST) = PyocsList().get_ocs_specific_list(xpath_id="Req_col_quantity",search_id=self.SearchID)
        (count, self.PACKAGE_LIST) = PyocsList().get_ocs_specific_list(xpath_id="Contract_col_package_type", search_id=self.SearchID)
        (count, self.PRODUCT_LIST) = PyocsList().get_ocs_specific_list(xpath_id="Contract_col_account_prod_name", search_id=self.SearchID)
        (count, self.COUNTRY_LIST) = PyocsList().get_ocs_specific_list(xpath_id="Req_col_SW_DefaultCountry", search_id=self.SearchID)

    def write_append(self,filename):

        wb = load_workbook(filename)
        sheet = wb.sheetnames

        for name in sheet:
            if name == self.projectID:
                ws = wb[name]
                rows = 1

                columndata = []
                for i in range(1, 5000):
                    cellvalue = ws.cell(row=i, column=1).value
                    if str(cellvalue) != "None":
                        columndata.append(cellvalue)
                        rows = rows + 1

                print("rows = " + str(rows))

                for i in range(0, len(self.OCS_LIST)):
                    ocs_value = str(self.OCS_LIST[i]).replace(u'\xa0', u' ')
                    customer_value = str(self.CUSTOMER_LIST[i]).replace(u'\xa0', u' ')
                    user_value = str(self.USER_LIST[i]).replace(u'\xa0', u' ')
                    data_value = str(self.DATA_LIST[i]).replace(u'\xa0', u' ')
                    number_value = str(self.NUMBER_LIST[i]).replace(u'\xa0', u' ')
                    package_value = str(self.PACKAGE_LIST[i]).replace(u'\xa0', u' ')
                    product_value = str(self.PRODUCT_LIST[i]).split(".")[0].replace(u'\xa0', u' ')
                    country_value = str(self.COUNTRY_LIST[i]).replace(u'\xa0', u' ')

                    print("\n-------------------------------------\n")
                    print("备注订单：" + ocs_value)
                    if ocs_value != " " and ocs_value != "无":
                        if ocs_value not in columndata:
                            ws.cell(row=rows, column=1).value = ocs_value
                            ws.cell(row=rows, column=2).value = customer_value
                            ws.cell(row=rows, column=3).value = user_value
                            ws.cell(row=rows, column=4).value = data_value
                            ws.cell(row=rows, column=5).value = number_value
                            ws.cell(row=rows, column=6).value = package_value
                            ws.cell(row=rows, column=7).value = product_value
                            ws.cell(row=rows, column=8).value = country_value
                            rows = rows + 1

                            print(ocs_value + " " + customer_value + " " + user_value + " " + data_value + " " + number_value + " " + package_value + " " + product_value + " " + country_value)

                        else:
                            index = columndata.index(ocs_value)
                            cellvalue = list(range(0, 8))
                            for j in range(0,8):
                                cellvalue[j] = str(ws.cell(row=index+1, column=j+1).value)

                            print("订单已存在excel中，不需要再次添加，仅作更新")
                            if customer_value != cellvalue[2]:
                                ws.cell(row=index+1, column=2).value = customer_value

                            if user_value != cellvalue[1]:
                                ws.cell(row=index+1, column=3).value = user_value

                            if data_value != cellvalue[3]:
                                ws.cell(row=index+1, column=4).value = data_value

                            if number_value != cellvalue[4]:
                                ws.cell(row=index+1, column=5).value = number_value

                            if package_value != cellvalue[5]:
                                ws.cell(row=index+1, column=6).value = package_value

                            if product_value != cellvalue[6]:
                                ws.cell(row=index+1, column=7).value = product_value

                            if country_value != cellvalue[7]:
                                ws.cell(row=index+1, column=8).value = country_value

                    else:
                        print("无效OCS号")

        wb.save(filename)

    def run(self):
        print(self.filepath)
        self.loadOcsData()
        self.write_append(self.filepath)

if __name__ == '__main__':
    SearchSW = OcsSoftware()
    SearchSW.run()



