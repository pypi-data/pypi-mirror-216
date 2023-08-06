from openpyxl import load_workbook

class OcsSoftware:
    SearchCusID = ""
    filename = ""

    def __init__(self):
        self.SearchType = "ALL"
        SearchID = "526913/F19500C"
        self.SearchOCSID = SearchID.split('/')[0]
        self.SearchCusID = SearchID.split('/')[1]
        self.filename = "save_ocs_data.xlsx"
        self.outputfile = "find_result.txt"

    def findCusID(self, type, OCSID ,CusID, filename):
        wb = load_workbook(filename)
        sheet = wb.sheetnames
        OutPutFile = open(self.outputfile, "w")

        needindex = 0
        if type != "ALL":
            ws = wb[sheet[0]]
            for index in range(0, 10):
                if type == str(ws.cell(row=1, column=index + 1).value):
                    needindex = index + 1

        for name in sheet:
            ws = wb[name]

            for i in range(2, 5000):
                if str(ws.cell(row=i, column=1).value) == "None":
                    break

                OCSvalue = str(ws.cell(row=i, column=1).value)
                Cusvalue = str(ws.cell(row=i, column=2).value)

                if OCSID == "0":
                    if CusID == Cusvalue:
                        checkflag = 1
                    else:
                        checkflag = 0
                elif CusID == "0":
                    if OCSID == OCSvalue:
                        checkflag = 1
                    else:
                        checkflag = 0
                else:
                    if CusID == Cusvalue and OCSID == OCSvalue:
                        checkflag = 1
                    else:
                        checkflag = 0


                if checkflag == 1:
                    outputstr = ""
                    if needindex != 0:
                        outputstr = outputstr + str(ws.cell(row=i, column=needindex).value) + " "
                    else:
                        for index in range(0,8):
                            outputstr = outputstr + str(ws.cell(row=i, column=index+1).value) + " "

                    outputstr = outputstr + '\n'
                    print(outputstr)
                    print(outputstr,file=OutPutFile)


        wb.save(filename)

    def run(self):
        self.findCusID(self.SearchType,self.SearchOCSID,self.SearchCusID,self.filename)

if __name__ == '__main__':
    SearchSW = OcsSoftware()
    SearchSW.run()



