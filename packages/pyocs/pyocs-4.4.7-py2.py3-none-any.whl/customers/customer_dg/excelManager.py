import openpyxl
from customers.customer_dg.cfgManager import CfgManager
from pyocs.pyocs_filesystem import PyocsFileSystem
from customers.customer_dg.cfgManager import work_dir
class ExcelManager(object):
    def __init__(self,excelName):
        self.work_book = openpyxl.load_workbook(excelName)
        self.work_sheet = self.work_book.active
        self.min_sign_ret = 0

    def get_rule_value(self, ruleList, min_row, max_row, max_col):
        '''拿到对应规则的值
        rulelist:规则的集合
        返回值是包含值的元组的列表'''
        if max_row < min_row:
            max_row = self.work_sheet.max_row
        valueList = list()
        value = list()
        col = 1
        while col <= max_col:
            row = min_row
            if self.work_sheet.cell(1,col).value in ruleList:
                value.clear()
                while row <= max_row:
                    value.append(self.work_sheet.cell(row,col).value)
                    row += 1
                valueList.append(list(value))
            col += 1
        # return valueList
        return valueList

    def find_all_index(self,src_list,value,index_list = None):
        if index_list is None:
            print("don't input the none index")
            return
        
        new_index = list()
        for i in index_list:
            if str(src_list[i]) == str(value) and str(src_list[i]) != "None":
                new_index.append(i)
        
        return new_index

    def transpose_list(self,src_list):
        return [[row[i] for row in src_list] for i in range(len(src_list[0]))]

    def match(self, list_src, list_dst):
        row = 0
        result = list()
        while row < len(list_src):
            col = 0
            index = list(range(len(list_dst[0])))
            while col < len(list_dst):
                index = self.find_all_index(list_dst[col],str(list_src[row][col]),index)
                col += 1 
                #print(index)
                if index == []:
                    break
            result.append(index)
            print("第 %s 行结果是：" % (row + 1))
            if index != []:
                print("有匹配项，行号是：")
                print(index)
            else:
                print("无匹配项或者存在关键项为空")
            row += 1
        return result

    @staticmethod
    def get_line_number_from_excel_for_changhon(file_path):
        changhon_url = "https://drive.cvte.com/p/Dc5ndx0Q-K8BGJv_Ag"
        rule_url = "https://drive.cvte.com/p/DfbbzKEQ-K8BGL-PAw"

        PyocsFileSystem.get_file_from_nut_driver(rule_url, work_dir)
        cfgm = CfgManager(work_dir + "/rule.xml")
        excelm = ExcelManager(file_path)
        PyocsFileSystem.get_file_from_nut_driver(changhon_url, work_dir)
        rule_list = cfgm.get_rule("CHANGHON")
        value_list = excelm.get_rule_value(rule_list, 2, excelm.work_sheet.max_row, excelm.work_sheet.max_column)
        transpose_value_list = excelm.transpose_list(value_list)
        return excelm.match(transpose_value_list,
                            ExcelManager(work_dir + '/KB信息对照表.xlsx').get_rule_value(rule_list, 2, excelm.work_sheet.max_row, excelm.work_sheet.max_column))

    def get_line_number_from_excel_for_TCL(self):
        ##get the sign line number
        ret = self.work_sheet.max_row
        while ret >= 1:
            if self.work_sheet.cell(ret, 1).value == "此处是标志2，请勿修改":
                max_sign_ret = ret
            elif self.work_sheet.cell(ret, 1).value == "此处是标志1，请勿修改":
                self.min_sign_ret = ret
                break
            ret -= 1
        #####
        cfgm = CfgManager(work_dir + "/rule.xml")
        rule_list = cfgm.get_rule("TCL")
        print(rule_list)
        value_list = self.get_rule_value(rule_list, self.min_sign_ret + 3, max_sign_ret - 1, 15)
        print(value_list)
        transpose_value_list = self.transpose_list(value_list)
        return self.match(transpose_value_list, self.get_rule_value(rule_list, 2, self.min_sign_ret - 2, 15))

    @staticmethod
    def get_ocs_number_from_excel(file_path, line_num):
        return ExcelManager(file_path).get_rule_value(['OCS'], 2, 1, 15)[0][line_num]

if __name__ == "__main__":
    m = ExcelManager("res/TCL_sheet.xlsx")
    #print(m.work_sheet.cell(1,1).value)
    m.get_line_number_from_excel_for_TCL()
    # result = ExcelManager.get_line_number_from_excel_for_changhon("CHANGHON", "customers/customer_dg/res/冻结书.xlsx")
    # for i in result:
    #     print(ExcelManager.get_ocs_number_from_excel(i[0]))

    # confirm_task = pyocs_software.PyocsSoftware()
    # result = confirm_task.get_ocs_number_from_abstract("[B.NEW][C.NEW][广州韩乐电器实业]T.MT5522.81D配HV650QUB-N9A（越南，POORD000091，客料号B.ZB.066.001")
    # print

    # PyocsFileSystem.get_file_from_nut_driver("https://drive.cvte.com/p/Dc5ndx0Q-K8BGJv_Ag","res/")
    # em = ExcelManager('res/冻结书.xlsx')
    # duizhaobiao = ExcelManager('res/KB信息对照表.xlsx')
    # #em = ExcelManager('KB信息对照表.xlsx')
    # rm = CfgManager('res/rule.xml')
    # rlist = rm.get_rule("CHANGHON")
    # print(rlist)
    # # rlist = rm.get_customer_list()
    # # print(rlist)
    # valueList = em.get_rule_value(rlist)
    # duizhaoxinxi = duizhaobiao.get_rule_value(rlist)
    # newlist = em.transpose_list(valueList)
    # print(newlist)
    # result = em.match(newlist,duizhaoxinxi)
    # print(result)