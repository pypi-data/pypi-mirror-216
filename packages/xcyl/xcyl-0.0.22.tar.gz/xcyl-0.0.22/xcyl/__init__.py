import os, sys, requests, json, pandas as pd, sqlite3, threading
from abc import abstractmethod
import datetime
from openpyxl import load_workbook
from copy import deepcopy as dc
from mystring import frame
from ephfile import ephfile
from pydbhub import dbhub


class flex(object):
	def __init__(self, enter_lambda, exit_lambda):
		self.enter_lambda = enter_lambda
		self.exit_lambda = exit_lambda

	def __enter__(self):
		self.enter_lambda()
		return self

	def __exit__(self, exc_type, exc_val, exc_tb):
		self.exit_lambda()
		return self

class core(object):
	current_table:frame=None
	table_name:str=None
	cur_data_sets = {}
	sub_lock:threading.Lock=threading.Lock()

	def __init__(self, echo:bool=False, lightload:bool=False):
		self.echo = echo
		self.lightload=lightload
		self.lock = flex(
			enter_lambda=lambda:self.sub_lock.acquire(),
			exit_lambda=lambda:self.sub_lock.release()
		)
		self.lload = flex(
			enter_lambda=lambda:self.__enter__() if self.lightload else None,
			exit_lambda=lambda:self.__exit__() if self.lightload else None
		)

	@abstractmethod
	def login(self):
		pass

	@abstractmethod
	def logout(self):
		pass

	@abstractmethod
	def open(self):
		pass

	@abstractmethod
	def close(self):
		pass

	@abstractmethod
	def tables(self):
		pass

	@abstractmethod
	def __internal_add_frame(self, table_name:str, database:pd.DataFrame):
		pass

	@abstractmethod
	def __internal_load_table(self, table_name:str):
		pass

	@abstractmethod
	def __internal_drop_table(self, table_name: str):
		pass

	def __iter__(self):
		return self.tables().__iter__()

	def __enter__(self):
		self.login()
		self.open()
		return self

	def __exit__(self, exc_type, exc_val, exc_tb):
		self.logout()
		self.close()
		return self

	def add_frame(self, dataset:object, dataname:str=None):
		if dataname is None:
			dataname="none"

		if isinstance(dataset, pd.DataFrame):
			self.__internal_add_frame(dataname, dataset)
		elif isinstance(dataset, frame):
			self.__internal_add_frame(dataname, dataset.df)

	def __iadd__(self, dataset:object):
		self.add_frame(dataset)
		return self

	def add_file(self, datapath:str):
		if os.path.isfile(datapath) or datapath.startswith("https://"):
			content = None
			if datapath.endswith(".csv"):
				content = pd.read_csv(datapath)
			elif datapath.endswith(".xlsx"):
				content = pd.read_excel(datapath)
			elif datapath.endswith(".zip"):
				content = pd.read_csv(datapath,encoding='ISO-8859-1')
			elif datapath.endswith(".txt"):
				content = pd.read_csv(datapath)
			elif datapath.endswith(".tsv"):
				content = pd.read_csv(datapath,seperator='	')
			elif datapath.startswith("https://"):
				content = pd.read_html(datapath)

			if content is not None:
				self.add_frame(content, os.path.basename(datapath))
		return

	def load_table(self, table_name:str):
		output = self.__internal_load_table(table_name)
		if not output.empty():
			self.cur_data_sets[table_name] = output
			self.current_table_name = table_name
		return output

	def drop_table(self):
		output = False
		if self.current_table_name:
			self.__internal_drop_table(self.current_table_name)
			self.current_table_name = None
		return output

	def current(self, table_name:str):
		if table_name in self:
			self.current_table = frame(self[table_name])
		else:
			print("Table not available")
	
	def load_add(self):
		for table_name in self.tables:
			self.cur_data_sets[table_name] = frame(self.load_table(table_name))

	def query(self, string:str):
		output = None
		if isinstance(self.current_table, pd.DataFrame):
			output = self.current_table.query(string)
		return output

	def __getitem__(self, item):
		return frame(self.tables[item])

	def __setitem__(self, key, value):
		self.add_frame(key, value)

	def __delitem__(self, key):
		if key in self.tables():
			del self.cur_data_sets[key]

	def __iter__(self):
		return self.cur_data_sets.__iter__()

	def __str__(self):
		return self.tables()

	def __contains__(self, item: str):
		return item in self.tables()

	def __call__(self, string: str):
		return self.query(string)

	def __len__(self):
		return len(self.tables())
	
	def csv(self, table_name:str = None):
		table_data = self[table_name] if table_name in list(self.table_names()) else None
		if table_data is not None:
			table_data.to_csv("{0}.csv".format(table_name))
	
	def sqlite(self):
		to_sql = sqlobj("temp.sqlite")
		for key, value in self.cur_data_sets.items():
			to_sql[key] = value
		return to_sql

class dbsite(core):
	def __init__(self, repo: str, access_token: str, owner: str, table_name: str):
		super().__init__()
		self.repo = repo
		self.access_token = access_token
		self.owner = owner
		self.config = "config.ini"
		self.table_name = table_name
		self.db = dbhub.Dbhub(config_file=self.config)
		self.current_db = frame(pd.DataFrame())
		self.current_db_name = None

	def login(self):
		pass

	def logout(self):
		pass

	def open(self):
		if os.path.exists(self.config):
			try:
				os.remove(self.config)
			except:
				pass

		with open(self.config, "w+") as config:
			config.write(f"""[dbhub]
	api_key = {self.access_token}
	db_owner = {self.owner}
	db_name = {self.repo}
	""")
		return self

	def close(self, exc_type, exc_val, exc_tb):
		self.set_currentdb(None,True)
		try:
			os.remove(self.config)
		except:
			pass
		return self

	def databases(self):
		files = []

		try:
			# https://github.com/LeMoussel/pydbhub/blob/5fac7fa1b136ccdac09c58165ab399603c32b16f/examples/list_databases/main.py#L28
			databases, err = self.db.Databases()
			if err is None:
				files = databases
		except:
			pass

		return files

	def set_currentdb(self,db_name:str=None, update_current:bool=False):
		if self.current_db_name is not None and update_current:
			self.updatedb()

		self.current_db = pd.DataFrame()
		self.current_db_name = None

		if db_name is not None and db_name in self.databases:
			self.current_db = self.load_table(db_name)
			self.current_db_name = db_name

		return self.current_db

	def updatedb(self):
		if self.currentdb_name == None:
			return

		with self.lock:
			with self.lload:
				with ephfile(self.current_db.to_sqlite()) as eph:
					try:
						db_contents = open(eph(), 'rb')
						with db_contents:
							# https://github.com/LeMoussel/pydbhub/blob/5fac7fa1b136ccdac09c58165ab399603c32b16f/examples/upload/main.py#L51
							res, err = self.db.Upload(db_name=eph(), db_bytes=db_contents,
												info=dbhub.UploadInformation(
													commitmsg=f"Uploading changes to {self.currend_db_name}",
													committimestamp=datetime.datetime.now(),
												))
							if err is not None:
								print(f"[ERROR] {err}")
								sys.exit(1)
					except Exception as e:
						pass
		return

	def tables(self):
		output = []

		if self.current_db is not None:
			try:
				# https://github.com/LeMoussel/pydbhub/blob/5fac7fa1b136ccdac09c58165ab399603c32b16f/examples/list_tables/main.py#L28
				tables, err = self.db.Tables(self.owner, self.self.current_db)
				if err is None:
					output = tables
			except:
				pass

		return output

	def __internal_add_frame(self, sheet_name, dataframe):
		with self.lock:
			while sheet_name in list(self.cur_data_sets.keys()):
				sheet_name += "_"

			self.cur_data_sets[sheet_name] = dataframe

	def delete_db(self, confirm:bool=False):
		with self.lock:
			with self.lload:
				if self.current_db and confirm:
					self.db.Delete(self.current_db)

	def __internal_drop_table(self):
		if self.table_name:
			with self.lock:
				with self.lload:
					self.query("DROP TABLE IF EXISTS {0}".format(self.table_name))

	def onquery(self, string: str):
		output = None

		try:
			results, err = self.db.Query(
				self.owner,
				self.repo,
				string.replace(':table_name', self.table_name).replace(':table_name'.upper(), self.table_name)
			)

			output = frame.from_arr(results)
		except:
			pass

		return output

	def __internal_load_table(self, table_name:str):
		output = pd.DataFrame()
		with self.lock:
			with self.lload:
				if self.lightload:
					output = self.onquery("SELECT * FROM {0}".format(table_name))
				else:
					# https://github.com/LeMoussel/pydbhub/blob/5fac7fa1b136ccdac09c58165ab399603c32b16f/examples/download_database/main.py#L29
					buf, err = self.db.Download(db_name=table_name, db_owner=self.owner)
					if err is not None:
						print(f"[ERROR] {err}")
					else:
						with open(table_name + ".sql", "wb") as sqlite_file:
							sqlite_file.write(buf)
					output = pd.read_csv(table_name + ".sql")

		return frame(output)

class xcylobj(core):
	"""
	the new excel object
	"""
	def __init__(self, filename: str = "TEMP_VALUE", useIndex: bool = False, useSheets: bool = True):
		super().__init__()
		if not filename.endswith(".xlsx"):
			filename += ".xlsx"
		self.filename = filename
		self.useIndex = useIndex
		self.useSheetNames = useSheets

	def login(self):
		pass

	def logout(self):
		pass

	def open(self):
		with self.lock:
			if os.path.exists(self.filename):
				for sheet_name in load_workbook(self.filename, read_only=True, keep_links=False).sheetnames:

					self.cur_data_sets[self.filename+"_"+sheet_name] = pd.read_excel(self.filename, engine="openpyxl", sheet_name=sheet_name)

	def close(self):
		with self.lock:
			if os.path.exists(self.filename):
				os.system("mv {} {}".format(self.filename, self.filename.replace(".xlsx", "_backup.xlsx")))

			try:
				keys, sheet_names, temp_name = list(self.cur_data_sets.keys()), [], "gen_name"
				for itr, key in enumerate(keys):
					og_sheet_name, final_sheet_name = key, key
					if len(key) > 20:
						self[temp_name+"_"+str(itr)] = dc(self[key])
						del self[key]
						final_sheet_name = temp_name+"_"+str(itr)

					sheet_names += [{
						"og_sheet_name": og_sheet_name,
						"final_sheet_name": final_sheet_name
					}]

				def dyct_frame(raw_dyct):
					dyct = dc(raw_dyct)
					for key in list(raw_dyct.keys()):
						dyct[key] = [dyct[key]]
					return pd.DataFrame.from_dict(dyct)

				if self.useSheetNames:
					self['sheet_names'] = pd.concat(list(map(
						dyct_frame,
						sheet_names
					)), ignore_index=True)

				with pd.ExcelWriter(self.filename, engine="xlsxwriter") as writer:
					for key, value in self.cur_data_sets.items():
						value.to_excel(writer, sheet_name=key, startrow=1, header=False, index=self.useIndex)
						worksheet = writer.sheets[key]
						(max_row, max_col) = value.shape
						worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': [{'header': column} for column in value.columns]})
						worksheet.set_column(0, max_col - 1, 12)
			except Exception as e:
				print("Exception :> {}".format(e))
				for key, value in self.cur_data_sets.items():
					value.to_csv(str(key) + ".csv")

	def __internal_add_frame(self, sheet_name, dataframe):
		with self.lock:
			while sheet_name in list(self.cur_data_sets.keys()):
				sheet_name += "_"

			self.cur_data_sets[sheet_name] = dataframe

	def tables(self):
		return self.cur_data_sets.keys()
	
	def __internal_load_table(self, table_name:str):
		output = pd.DataFrame()
		with self.lock:
			with self.lload:
				output = self.cur_data_sets[table_name]
		return output

	def __internal_drop_table(self, table_name):
		output = False
		with self.lock:
			with self.lload:
					self.cur_data_sets.pop(table_name)
		return output

class sqlobj(core):
	"""
	Sample usage:
	```
	with sqlobj("dataset.sqlite") as sql:
		container = pd.read_sql(sql.table_name, sql.connection_string)
	...
	with sqlobj("dataset.sqlite", threadLock=<x>) as sql:
		container = pd.read_sql(sql.table_name, sql.connection_string)
	...
	with sqlobj("dataset.sqlite") as sql:
		container.to_sql(sql.table_name, sql.connection, if_exists='replace')
	```
	"""

	def __init__(self, file_name: str):
		super().__init__()
		self.file_name = file_name
		self.table_name = "dataset"
		self.connection = None

	def login(self):
		pass

	def logout(self):
		pass

	def open(self):
		with self.lock:
			self.connection = sqlite3.connect(self.file_name)
			if not self.lightload:
				current_cursor = self.connection.cursor()
				current_cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
				for name in current_cursor.fetchall():
					self.cur_data_sets[name[0]] = pd.read_sql_query("""SELECT * FROM "{0}";""".format(name[0]), self.connection)
				current_cursor = None

	def close(self):
		with self.lock:
			self.connection.close()
		return self

	def tables(self):
		output = []

		with self.lload:
		
			if self.lightload:
				current_cursor = self.connection.cursor()
				current_cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")

				for name in current_cursor.fetchall():
					self.cur_data_sets += [name[0]]

				current_cursor = None
			else:
				tables = self.cur_data_sets.keys()
		
		return output

	def __internal_load_table(self, table_name):
		output = pd.DataFrame()

		with self.lload:

			if self.lightload:
				output = pd.read_sql_query("""SELECT * FROM "{0}";""".format(table_name), self.connection)
			else:
				output = self.cur_data_sets[table_name]

		return output

	def __internal_add_frame(self, sheet_name, dataframe):
		while sheet_name in list(self.tables()):
			sheet_name += "_"

		with self.lock:
			with self.lload:

				current_cursor = self.connection.cursor()
				dataframe.to_sql(sheet_name, self.connection, if_exists='replace')
				current_cursor = None

				if not self.lightload:
					self.cur_data_sets[sheet_name] = dataframe

		return

	def __internal_drop_table(self, table_name, backup=True):
		with self.load:
			with self.lload:

				if table_name in self.table_names:
					if backup:
						self[table_name].to_csv(".backup_{0}.csv".format(table_name))

					if self.lightload:
						current_cursor = self.connection.cursor()
						current_cursor.execute("""DROP table IF EXISTS "{0}";""".format(table_name))
						current_cursor = None
					else:
						try:
							self.cur_data_sets.pop(table_name)
						except: pass

		return