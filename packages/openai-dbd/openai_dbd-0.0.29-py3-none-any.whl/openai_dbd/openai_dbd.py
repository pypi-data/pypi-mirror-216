import string, random, os, requests, base64, zipfile
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side

def callOpenAI(URL, HEADERS, MY_PROMPT):
    OpenAIData = requests.post(URL, headers = HEADERS, json = {
        "model": 'gpt-3.5-turbo',
        "messages": [
            {
            "role": "user",
            "content": MY_PROMPT
            }
        ],
        "temperature": 0.7
    })

    return OpenAIData.json()

def giveRandName(l = 7):
    res = ''.join(random.choices(string.ascii_uppercase +
                             string.digits, k=l))
    return str(res)

def sql_prompt(projectCategory, featureOfDBStructure, open_ai_data):
    promptX = f"""
    Generate SQL Queries for *{projectCategory}* that provides features like
    `{featureOfDBStructure}`

    Maintain proper relations between these tables and provide information for each table. 
    Additionally, normalize the tables for better understanding and performance, and add an 'id' field with 'auto_increment' to each table.
    Add created_at and updated_at fields to each table.
    Place all root tables at the top, as other tables depend on them.
    I want table_name, column_name, data_type, size and constraint.
    In the constraint column, capitalize all letters except for the table names.
    Keep tables name small everywhere
    Keep in mind that dont use keywords that are researved for mysql

    I want you to only reply the sql command and nothing else, do not write explanations and any starting and ending sentences follow my instructions strictly.
    """

    return callOpenAI(open_ai_data['url'], open_ai_data['headers'], promptX)

def convert_json_to_sql(json_obj, open_ai_data):
    try:
        promptX = f"""
        Hi, GPT 
        I want you to act as json to sql converter and i want you to strictly follow my instructions. 
        I will give you json format and you will only return me only and only sql statments.

        Here is my Json Format
        `{json_obj}`
        Only response in Pure SQL format, 
        Don't include any kind of extra text at the starting and ending of your response and dont add any kind of explanations.
        """

        return callOpenAI(open_ai_data['url'], open_ai_data['headers'], promptX)
    except Exception as ex:
        return {"success": False, "message": str(ex)}

def json_prompt(projectCategory, featureOfDBStructure, open_ai_data):
    promptX = """
    Generate Database Architecture for """ + projectCategory + """ that provides features like
    """ + featureOfDBStructure + """
    Maintain proper relations between these tables and provide information for each table. Additionally, normalize the tables for better understanding and performance, and add an 'id' field with 'auto_increment' to each table.
    Add created_at and updated_at fields to each table.
    Place all root tables at the top, as other tables depend on them.
    I want table_name, column_name, data_type, size and constraint.
    In the constraint column, capitalize all letters except for the table names.
    Keep in mind that dont use keywords that are researved for mysql

    `Do not include any explanations and extra content`, only provide formatted JSON response

    Json response example
    [{
        "table_name": "Table Name",
        "columns": [{"column_name": "Col Name", "data_type": "int", "size": "15", "constraint": "constraints"}, {"column_name": "Col Name", "data_type": "int", "size": "15", "constraint": "constraints"}]
    },
    {
        "table_name": "Table Name",
        "columns": [{"column_name": "Col Name", "data_type": "int", "size": "15", "constraint": "constraints"}, {"column_name": "Col Name", "data_type": "int", "size": "15", "constraint": "constraints"}]
    }]
    """

    return callOpenAI(open_ai_data['url'], open_ai_data['headers'], promptX)

def correct_schema(features = '', sql_schema = '', open_ai_data = {}):
    promptX = f"""
    I am learning SQL and i want your help, 
    I have designed database schema but i am not sure weather it is right or not.
    I want you to act as SQL Expert/SQL Translator and Correct my mistakes and help me to learn SQL.

    I want your help to look into my schema, and help me to correct all tables and normalization and modify table details,
    If you find it may cause performance issue you can modify it.
    Also i want you to check constraints in it, if you find something missing you can add it.
    If you feel any table fields are required to added, you can add them as well.
    Also add my created tables in response with modifications you have done.

    `{"Also these are some features i want you to add to existing schema and that are as below" + features if features != '' else ''}`

    Here is my SQL schema
    `{sql_schema}`

    I want you to only response with `Pure SQL queries`,
    Don't give explanations or any text,
    Also don't add starting or ending text in response which you usually write.
    """

    return callOpenAI(open_ai_data['url'], open_ai_data['headers'], promptX)

def generate_sql(db_name, data):
    fName = giveRandName(10)
    with open(f"{fName}.sql", 'w') as file:
        dbName = db_name.replace(' ', '_').replace('-', '_').lower()
        file.write(f"CREATE DATABASE {dbName};\n\n")
        file.write(f"USE {dbName};\n\n")
        file.write(data)

    response = HttpResponse(open(f'{fName}.sql', 'rb'))
    response['Content-Disposition'] = f'attachment; filename={fName}.sql'
    os.remove(f'{fName}.sql')
    return response

def generate_excel(file_name, json_data):
    wb = Workbook()
    sheet = wb.active

    bold_font = Font(bold=True)
    table_name_font = Font(bold=True, size=16)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    column_widths = {
        "column_name": 0.20,
        "data_type": 0.15,
        "size": 0.1,
        "constraint": 0.45
    }

    for index, table in enumerate(json_data):
        table_start_row = sheet.max_row + 1
        table_end_row = table_start_row + 1
        sheet.merge_cells(start_row=table_start_row, start_column=3, end_row=table_end_row, end_column=6)
        cell = sheet.cell(row=table_start_row, column=3, value=f"{index + 1}. {table['table_name']}")
        cell.font = table_name_font
        # cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.alignment = Alignment(vertical="center")

        headers = ["column_name", "data_type", "size", "constraint"]
        sheet.append([""] * 2 + [header.upper() for header in headers])

        sheet.append([""] * 2 + [header.upper() for header in headers])
        for col, header in enumerate(sheet[sheet.max_row], start=3):
            header.font = bold_font

        for column in table["columns"]:
            column_name = column["column_name"]
            data_type = column["data_type"].upper()
            size = column["size"]
            constraint = column["constraint"]

            sheet.append([""] * 2 + [column_name, data_type, size, constraint])
        sheet.append([""] * 6 )
        
        table_start_cell = sheet.cell(row=table_start_row, column=3)
        table_end_cell = sheet.cell(row=sheet.max_row, column=6)
        table_range = table_start_cell.coordinate + ':' + table_end_cell.coordinate
        for row in sheet[table_range]:
            for cell in row:
                cell.border = thin_border
                
        for col in range(3, sheet.max_column + 1):
            column_name = sheet.cell(row=table_start_row + 1, column=col).value
            width = column_widths.get(column_name.lower(), 1)
            sheet.column_dimensions[chr(col + 64)].width = width * 100
            
        sheet.append([])
        empty_row = sheet.max_row
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=empty_row, column=col).border = None

    fName = giveRandName(10)
    file_name = file_name.replace(' ', '_').replace('-', '_').lower()
    wb.save(f'{fName}.xlsx')
    data = open(f'{fName}.xlsx', 'rb').read()
    base64_encoded = base64.b64encode(data).decode('UTF-8')
    os.remove(f'{fName}.xlsx')
    return HttpResponse(base64_encoded)

    # response = HttpResponse(open(f'{fName}.xlsx', 'rb'))
    # response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
    # os.remove(f'{fName}.xlsx')
    # return response

def generate_django_models(table_json):
    model_files = []

    for table in table_json:
        table_name = table['table_name']
        columns = table['columns']

        model_content = "from django.db import models\n\n"
        model_content += f"class {table_name.capitalize()}(models.Model):\n"

        for column in columns:
            column_name = column['column_name']
            data_type = column['data_type']
            size = column['size']
            constraint = column['constraint']

            if(data_type.lower() == "int" or data_type.lower() == "integer"):
                if(constraint):
                    if "FOREIGN KEY" in constraint:
                        constraint = constraint.split("REFERENCES")
                        reference_table = constraint[1].split("(")[0].strip()
                        reference_column = constraint[1].split("(")[1].split(")")[0].strip()
                        model_content += f"    {column_name} = models.ForeignKey('{reference_table}', on_delete=models.CASCADE, related_name='{table_name}_set')"
                        model_content += "\n"
                        continue
                    else:
                        model_content += f"    {column_name} = models.IntegerField("
            elif(data_type.lower() == "text"):
                model_content += f"    {column_name} = models.TextField("
            elif(data_type.lower() == "varchar"):
                model_content += f"    {column_name} = models.CharField("
            elif(data_type.lower() == "timestamp"):
                model_content += f"    {column_name} = models.DateTimeField("
            else:
                model_content += f"    {column_name} = models.{data_type.capitalize()}Field("

            if size:
                model_content += f"max_length={size},"
            else:
                model_content += ""

            if constraint:
                if "PRIMARY KEY" in constraint:
                    model_content += "primary_key=True,auto_created=True,"
                elif "NOT NULL" in constraint:
                    model_content += "null=False,"
                elif "FOREIGN KEY" in constraint:
                    constraint = constraint.split("REFERENCES")
                    reference_table = constraint[1].split("(")[0].strip()
                    reference_column = constraint[1].split("(")[1].split(")")[0].strip()
                    model_content += f"(models.ForeignKey('{reference_table}', on_delete=models.CASCADE, related_name='{table_name}_set'))"
            
            model_content += ")"
            if(model_content[len(model_content) - 2:] == ",)"):
                model_content = model_content.replace(model_content[len(model_content) - 2:], ')')
            model_content += "\n"

        model_files.append({
            "filename": f"{table_name.lower()}.py",
            "file_content": model_content
        })

    return model_files

def generate_sequelize_models(table_json):
    model_files = []

    for table in table_json:
        table_name = table['table_name']
        columns = table['columns']

        model_content = "const { Sequelize } = require('sequelize');\n\n"
        model_content += f"const {table_name.capitalize()} = sequelize.define('{table_name}', {{\n"

        for column in columns:
            column_name = column['column_name']
            data_type = column['data_type']
            size = column['size']
            constraint = column['constraint']

            model_content += f"  {column_name}: {{\n"
            model_content += f"    type: Sequelize.{data_type.upper()}"
            
            if size:
                model_content += f"({size})"

            if constraint:
                model_content += f",\n    allowNull: false"
                if "FOREIGN KEY" in constraint:
                    constraint = constraint.split("REFERENCES")
                    reference_table = constraint[1].split("(")[0].strip()
                    reference_column = constraint[1].split("(")[1].split(")")[0].strip()
                    model_content += f",\n    references: {{\n      model: {reference_table.capitalize()},\n      key: '{reference_column}'\n    }}"
                
            model_content += "\n  },\n"
        
        model_content += "}, {\n"
        model_content += "  timestamps: false\n"
        model_content += "});\n\n"

        model_content += f"module.exports = {table_name.capitalize()};\n"

        model_files.append({
            "filename": f"{table_name.lower()}.js",
            "file_content": model_content
        })

    return model_files

def json_to_python_models(table_json):
    django_models = generate_django_models(table_json)

    zip_filename = f'{giveRandName(10)}.zip'
    with zipfile.ZipFile(zip_filename, 'w') as zipf:
        for model in django_models:
            filename = model['filename']
            file_content = model['file_content']
            zipf.writestr(filename, file_content)

    data = open(zip_filename, 'rb').read()
    base64_encoded = base64.b64encode(data).decode('UTF-8')
    os.remove(zip_filename)
    return HttpResponse(base64_encoded)

def json_to_sequelize_models(table_json):
    sequelize_models = generate_sequelize_models(table_json)

    zip_filename = f'{giveRandName(10)}.zip'
    with zipfile.ZipFile(zip_filename, 'w') as zipf:
        for model in sequelize_models:
            filename = model['filename']
            file_content = model['file_content']
            zipf.writestr(filename, file_content)
    
    # with open(zip_filename, "rb") as f:
    #     response = HttpResponse(f.read(), content_type="application/zip")
    #     response["Content-Disposition"] = f'attachment; filename="{zip_filename}"'

    # os.remove(zip_filename)
    # return response

    data = open(zip_filename, 'rb').read()
    base64_encoded = base64.b64encode(data).decode('UTF-8')
    os.remove(zip_filename)
    return HttpResponse(base64_encoded)
